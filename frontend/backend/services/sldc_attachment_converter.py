from __future__ import annotations

import base64
import csv
import io
import math
import re
from dataclasses import dataclass
from typing import Any, List, Optional, Sequence, Tuple


# Base XLSX template used by the frontend (`/templates/telangana_sldc_template.xlsx`).
# Embedded here so the backend Docker image (built from `backend/`) can generate
# identical Telangana SLDC attachments for cron-driven auto emails.
_TELANGANA_TEMPLATE_XLSX_B64 = """
UEsDBBQAAAAIAP1geFz81L63YwEAAG0FAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbK2Uy07DMBBF
93xF5C1K3LJACDXtgsISKlE+wNjTxqpfsqevv2eStLykBqp0Eyty7j3jO+OMJjtrsg3EpL0r2bAY
sAyc9Eq7Zcne5k/5HcsSCqeE8Q5KtofEJuOr0XwfIGUkdqlkFWK45zzJCqxIhQ/gaGfhoxVIr3HJ
g5ArsQR+MxjccukdgsMcaw82Hr0QP2oF2UxEfBaWMHxnOJIbtM9hQX4se2iFNbtkIgSjpUAqnG+c
+kXN/WKhJSgv15YkRWNzXbvwk8CEewOpNyqFCEKlCgCtKVrTI3kKC7E2mD3uyL3NPIJJ5/EOYRak
bL5JlQ5dhO4DndZtegZB+mkUWxqljtC3Pq7evV9dOvZ6LazQrqvpJJ5FHxInVO8CoE5OgcoDWUJE
/dX2Trb0Ec6HH4egVp9NXCf0tveBW5t/wg+tbvqSeLP0v9U/e/7p/9c9r0QE9YqRJvPi1/2797EO
3vwtxx9QSwMEFAAAAAgA/WB4XKRSuaX2AAAA3gIAAAsAAABfcmVscy8ucmVsc62Sz04DIRCH7z4F
4d5lW40xpmwvjUlvxqwPMMLsn+zCEJjq9u3l0ugmde2hR+A3H98MbHeTG8UnxtST13JdlFKgN2R7
32r5Xr+snqRIDN7CSB61PGGSu+pu+4YjcK5JXR+SyBCftOyYw7NSyXToIBUU0OeThqIDzsvYqgBm
gBbVpiwfVfzNkNWMKQ5Wy3iwD1LUp4DXsKlpeoN7MkeHni9cocwxMblViLk6cp9bETXEFllLS+Y1
b58zRWZLdVnp/pZKODF6i3ZZCkJYMtpcb/T3AyiHDBYYlKGI/wwpJ5aE1rcc0TzxYzON6ovi8EE0
nF3U7FtW31BLAwQUAAAACAD9YHhchCSxVukAAAC5AgAAGgAAAHhsL19yZWxzL3dvcmtib29rLnht
bC5yZWxzrZLBasMwEETv/Qqx91p2WkopkXMpgVxb9wOEtLZMbEloN23991UbSBwIoQefxKzYmcdI
6833OIhPTNQHr6AqShDoTbC97xR8NNv7ZxDE2ls9BI8KJiTY1HfrNxw05x1yfSSRTTwpcMzxRUoy
DkdNRYjo800b0qg5y9TJqM1edyhXZfkk09wD6gtPsbMK0s5WIJop4n+8Q9v2Bl+DOYzo+UqEJJ6G
zC8anTpkBUddZB+Q1+NXS8Zz3sVz+p88DqtbDA+LVuB0QvvOKT/wvIn5+BbM45IwXyHtySHyGeQ0
+kXNx6kZefHj6h9QSwMEFAAAAAgA/WB4XIaTPmqBAQAAJQMAABAAAABkb2NQcm9wcy9hcHAueG1s
nVLBbtswDL3vKwzdGznBMAyBrGJIN/SwYgGSdmdOpmOhsiSIjJHs6yc7iOusO82nx8eHp2eS6v7U
uaLHRDb4SiwXpSjQm1Bbf6jE8/7b3WdREIOvwQWPlTgjiXv9QW1TiJjYIhXZwVMlWua4lpJMix3Q
Ird97jQhdcC5TAcZmsYafAjm2KFnuSrLTxJPjL7G+i5OhuLiuO75f03rYIZ89LI/x+yn1ZcYnTXA
+Sf1kzUpUGi4+Hoy6JScN1U22qE5JstnXSo5L9XOgMNNNtYNOEIl3wj1iDDMbAs2kVY9r3s0HFJB
9nee2koUv4BwiFOJHpIFz+IiuxQjdpE46Z8hvVKLyKTkRI5wrp1j+1EvR0EGt0I5Bcn4NuLeskP6
0Wwh8T8SL+eJxwxilnGX91EfHb5LeH3rL/dN6CL4PEI5oSfwcMBBO6Hv1r/Sc9yHB2C8zviWVLsW
EtZ5LdMOJkI95rDJDfpNC/6A9VXzvjFcxMvl6vVytSjzNx7ClVPy7cD1H1BLAwQUAAAACAD9YHhc
5gxRVm0BAADjAgAAEQAAAGRvY1Byb3BzL2NvcmUueG1sfZLJbsIwEIbvfYrI1yrYCVC1UQjqIk5F
qlSq9uraA7jEi2zTwNvXWQiLUC+WZ/7fn2dGk093sox+wTqh1QQlA4IiUExzoVYT9LGYxfcocp4q
TkutYIL24NC0uMmZyZi28Ga1AesFuCiAlMuYmaC19ybD2LE1SOoGwaGCuNRWUh9Cu8KGsg1dAU4J
ucMSPOXUU1wDY9MTUYfkrEearS0bAGcYSpCgvMPJIMFHrwcr3dUHjXLilMLvDVy1HsTevXOiN1ZV
NaiGjTXUn+Cv+et702osVD0qBqjIOcuYBeq1LR4lVTk+SdTDK6nz8zDmpQD+tC9yfCXXNdO+Ax6F
IrK25IPyOXx+WcxQkZKUxOQhTpIFGWfjNCPklpBw1v+eQY5U2f30P3YUkzRO7xfJXTYK5NEl9kBp
OvbCl1A0vXbXcOPgmBXGhwVrpbNEiN32+weYb8U+CPPYwL7Slrt2Oseo3r3Qy0rbbnAX0dleFn9Q
SwMEFAAAAAgA/WB4XAXRl3BoAQAA7AMAABMAAABkb2NQcm9wcy9jdXN0b20ueG1svZNdT4MwFIbv
/RWk90ALg8ECLFth0WiiyT7uK5SNCC1py3Qx/neLGzPeGT92eXJOnudtTxtNX5ra2FMhK85igCwI
DMpyXlRsG4P1amEGwJCKsILUnNEYHKgE0+QqehC8pUJVVBqawGQMdkq1E9uW+Y42RFq6zXSn5KIh
Spdia/OyrHKa8rxrKFO2A6Fv551UvDHbMw4ceZO9+imy4HmfTm5Wh1bzkugEPxhlo6oiBq+ph9PU
g57pZCE2EURzM3TDsQkDCJ25gxfhLHsDRtsPO8BgpNEnv9Y0UVfsSeIdYVtaaPJeTR45r5OS1JJG
9lBG9qD8pdwd5Dd4c9TV7bNUIvHnmTv2/YU7grOR56BgFMAU+hgjP82yBf7Icpr9szSjIc3t8l7f
cNHlat5VdbGh4ks4BF3XRMhyLGghhFz4L2m8Ic1dv5N1u+IpUfQCS/EH8TInNcUadQHp+CzdEdE/
+O857c9/mrwDUEsDBBQAAAAIAP1geFxzkXtZuwUAAKYbAAATAAAAeGwvdGhlbWUvdGhlbWUxLnht
bO1ZT2/bNhS/71MQureybCl1gjpF7NjtlqYNErdDj7RES6wpUSDppL4N7XHAgGHdsMuA3XYYthVo
gV26T5Otw9YB/Qp7+mObiqk0aTNsQ+uDLZK/95/v8VG+eu1BzNAhEZLypGM5lxsWIonPA5qEHevO
cHCpbSGpcBJgxhPSsWZEWtc2P7iKN1REYoKAPJEbuGNFSqUbti19mMbyMk9JAmtjLmKsYChCOxD4
CNjGzG42Gmt2jGlioQTHwPX2eEx9goYZS2tzzrzP4CtRMpvwmTjwc4k6RY4NJk72I2eyxwQ6xKxj
gZyAHw3JA2UhhqWChY7VyD+WvXnVXhAxVUOr0Q3yT0lXEgSTZk4nwtGC0Bm461e2F/ybBf9VXL/f
7/WdBb8cgH0fLHVWsO6g7XTnPDVQ8bjKu9fwGm4Vr/FvreDXu92ut17Bt5Z4dwXfbqy5W80K3l3i
vVX9u1u93loF7y3xayv4wZX1NbeKz0ERo8lkBZ3FcxGZBWTM2Q0jvA3w9nwDLFG2trsK+kTV7bUY
3+diAIA8uFjRBKlZSsbYB1wPxyNBcSYAbxCsrRRTvlyZymQh6Quaqo71UYohI5aQV89/ePX8KXr1
/Mnxw2fHD38+fvTo+OFPBsIbOAl1wpffff7XN5+gP59++/Lxl2a81PG//fjpr798YQYqHfjiqye/
P3vy4uvP/vj+sQG+JfBIhw9pTCS6RY7QPo/BNoMAMhLnoxhGmFYocARIA7Cvogrw1gwzE65Lqs67
K6AAmIDXp/cruh5EYqqoAbgTxRXgLuesy4XRnJ1Mlm7ONAnNwsVUx+1jfGiS3TsR2v40hZ1MTSx7
Eamouccg2jgkCVEoW+MTQgxk9yit+HWX+oJLPlboHkVdTI0uGdKRMhPdoDHEZWZSEEJd8c3uXdTl
zMR+mxxWkZAQmJlYElZx43U8VTg2aoxjpiNvYhWZlDyYCb/icKkg0iFhHPUDIqWJ5raYVdTdwVCJ
jGHfZbO4ihSKTkzIm5hzHbnNJ70Ix6lRZ5pEOvZDOYEtitEeV0YleDVDsjHEASe14b5LiTpfWt+h
YWTeINnKVJhSgvBqPs7YGJOkrO+VSh3T5LSyzSjU7fdlew7fgkPMlDwni3Ud7n9YorfxNNkjkBXv
K/T7Cv0uVui6XL74urwsxbbea+ds4trGe0wZO1AzRm7KvIhLMC8YwGQ+yIkWfX4awWMproILBc6f
keDqY6qigwinIMbJJYSyZB1KlHIJtwurlnd+RaVgcz7nze+VgMZqlwfFdEu/by7Y5KNQ6oJaGYOz
CmtdeTthTgE8ozTHM0vzTpVma96EvEE4e5ngrDUL0bBRMCNB5veCwTwsFx4iGeGAlDFyjIY4rTO6
rf16r2nS1ltvJ+0sQdLFuTXivAuIUmMlSvZqOrKkOkJHoJXX9Czk47RjjaHngsc4BX4yK1WYhUnH
8lVpymuT+aTB5m3pNGoNrohIhVTbWEYFVb40fx2TLPVvem7mh4sxwFCNzqZFq+38i1rYJ0NLxmPi
q5qZ5bBc41NFxEEUHKERm4p9DHq7xe4KqISjojkfCMhQt9x41cwvs+Dka58yOzBLI1zWpLYW+wKe
Py90yEeaenaN7m9oSusCTfHeXVOynQsNbivIr17QBgiMsj3asbhQEYcqlEbUHwhoHHJZoBeCtMhU
Qix7h53pSg6XdavgURS5MFL7NESCQqVTkSBkT5V2voaZ09TP1zmjss4s1JVp8Tsih4QNs+xdy+y3
UDSvJqUjctzJoNmm7BqFg/9w5+PWdD6ntwdLQe55ehFXK/raUbD+diqc86htmi1uemc+alO4pqDs
Cwo3FT5b9rdDvg/RR4uOEsFGvNQu028xOQKd25pxGat/to1ahqBdE++LbD41Z7dqnH26uDd3tmfw
tXe6q+3VFLW1i0w+Wvkzi4/ug+xtuB9NmZLFe6cHcCntzf+GAD72knTzb1BLAwQUAAAACAD9YHhc
AsQwdCEDAAANDQAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1sdddNb9NAGATgO7/C8gkkgr3rz1Rp
UAmCS1sqUsR5lWwbC8dr7E1F/j37DlzIDrd0xtvHtaORunr/69gnL3aaOzdcp+pdniZ22Ll9Nzxf
p98ePy3aNJm9Gfamd4O9Ts92Tt+vX63m2Sfh6DBfpwfvx6ssm3cHezTzOzfaITRPbjoaH36cnrN5
nKzZzwdr/bHPdJ7X2dF0Q5qchu7nyW7cafABV026Xs3deuXX9+ZoE/eUfLaDnYx30yrz61Um5Z8L
tuEGzWSH7rJ46M3gkyGc/++RZOP6vrNTZ+fw8Tia4ZzcdsfO2/1ic7D9eIq4jRnNrvPn13ff31x2
H42PLF0u8mKhc11fNo/nMbq6G/xk9uYcqU6KnU/YoTvvzGV2M46TezF9cu8id/v49eZ+u/mSaRVe
gQ53V0bXuOSb7/rwd0bPbrO5ff3h4XaR58Xb5M+H+u8H3UaP5MH4w2X2oXe7H5HYhff8EF6F219W
n9xkd2b27JHfvGxYvPXGh+9xsg1fxf2pjx5Znl/lebjzK1WRSlVSFTmpCpwq2akynFLh10aVgqWY
pWApZilYilkKlmaWhqWZpWFpZmlYmlkaVsGsAlbBrAJWwawCVsGsAlbJrBJWyawSVsmsElbJrBJW
xawKVsWsClbFrApWxawKVs2sGlbNrBpWzawaVs2sGlbDrAZWw6wGVsOsBlbDrAZWy6wWVsusFlbL
rBZWy6wW1pJZS1hLZi1hLZm1hLVkloQLlRMLoVSxhVCq2EIoVWwhXCi2GwilIhZ2Q7HdQCgVsbAb
iu0GQqmIhd1QbDcQSkUs7IZiu4FQKmJhNxTbDYRSEQu7odhuIJSKWNgNxXYDoVTEwm4othsIpSIW
dkOx3UAoFbGwG4rtBkKpiIXdUGw3EEpFLOyGYruBUCpiYTcU2w2EUhELu6HYbiCUiljYDcV2A6FU
xMJuKLYbCKUiFnZDsd1AKBWxsBua7QZCqWILoVSxhVCq2EK40Gw3EEpFLOyGZruBUCpiYTc02w2E
UhELu6HZbiCUiljYDc12A6FUxMJuaLYbCKUiFnZD/7sbWfh/Zv0bUEsDBBQAAAAIAP1geFyGAyHl
7QMAALE5AAANAAAAeGwvc3R5bGVzLnhtbO2bX2+bMBDA3/cpLL9vhLRrtwmYtkyR9rBp2jJprwZM
sGpsZJwu2aefjQl/1naFZG3NAg/FPo67y69n+2KI93abUXCNRUE486H7YgYBZhGPCVv78Ptq+fwV
BIVELEaUM+zDHS7g2+CZV8gdxd9SjCVQFt4UOYrU1VzgAotrDLWQFT5MpczfOE4RpThDxQueY6au
JFxkSKquWDuFugfFhbaUUWc+m104GSIMBh7bZMtMFiDiGyZVbLUImNPHWAkvziEw5hY8VhHEsZNl
zk4d0Ak8p7IReAlnjalzaASBF4JrRH0408qk1S6kIFe4JdiYNlMQyuu/TN91dS/ilAsg1qEPl8tZ
eWgxQ1llYoEoCQUpQzKOO+7dsbqf93ePxO4u/4fTP9h/edJpQSit0+IMGkHg5UhKLNhSdUDVXu1y
3ETgGL17tNcC7dz5S212vWhHaQ5tKOxeaMJ3Wlb7+is4JfFNb+/PF5cfFrd4W6jDhHGrt/KkIIVc
xGqGqAei5mRkzr4ReBQnEpSzgg9lWg7gO/4xjlYNPEHWad87St3AkzzveYPSVH/3se3PHZcpImKQ
y3tuuN3lY2D5++c80MZjBB5yKXnW8x6j/G8+7QM47pdO/yL4Y0eBhSP1yRLhSeJ8EqfHWakaajGI
MKXftJEfSac02yatsmymizJWN9UyUjWNGdPRjtrWjO2W2bNLY3eb3LTZ9XW7A4DynO6WXNva9z5v
shCLZVkxtnSUgab3vrTR9N9RsmYZbpv5IrjEkSyLZik2WAWK9log5YL8Ul517bHGDAtEoa6xJYm0
yMCFQOKt/MolMlaU5Z8C5Ssl9GGCaKGK6CIVhF2t+JLUMgUtr30DyqMrHFchgJTEMWYtTWebTAD/
M4DuUICREmAxkJ/h8ej45m187mB87h/43AfBt5dYmH4dfnNL+alFzX54Z5bCG0vynVvKz6rkM7tX
h+JzTx3fXbn30tLcs23s3p1+F1P6HV72XR5d9p1G+k1l84PMfq+m2W9Kv6nwm9LvVNJvFPh6pJ8N
o7dP4Xeio9ftw+91h5878Wv4zXvt+c0mgEcl4ATwWIDuBPCoFfgPgPZ8/x1JBWMtP3tLmHHwO+n8
mw3lN+LnvsOfHVkzgB+a31kfft0thD4r8Mnw65V/w7dgrJkAreBn6x7CWPjZOv+Nhd/w55cnM//1
Wj8upvXjqPwbvgk4jd9OBT38zasJ4BgmwLHws7WAeeQJ0KneJ2+9tN55Zb2WAv1bMR9+1mjUN84a
erghVBJmep3X1ZXNeNu8qV5elSikuPIS4wRtqFzVIh827U84JpvsNdxrfSHXXFZaZbtRLX+a5zQ/
tgx+A1BLAwQUAAAACAD9YHhc7hj9ea4BAAD0AgAADwAAAHhsL3dvcmtib29rLnhtbI2SS28TMRDH
73yKle90t6hBEGVTiZZAJV4ioZwn9mx3VL/k8SZtPz1jJ4HCib3Y85/3z7u4fHC22WFiCr5X52ed
atDrYMjf9erHZvXyjWo4gzdgg8dePSKry+WLxT6k+20I943kzzmCFl9MyJh2qIrouVdjznHetqxH
dMBnIaIXzxCSgyxmumtZcsDwiJidbV913evWAfljhXn6nxphGEjjddCTQ58PRRJayLISjxRZLRcD
Wbw9bNlAjF/AybwPVjUWOL83lNH06kLMsMe/hDTFdxPZYsy6mWqXvzf/lhrhhIdSm5H459FRgsp5
S7jnP/HFbEBn2uEGtr0S0jDlsCKbMV1Dxg8pTLFyz2kSiAMlzuuCpgY78uToqcw1gGUJ4DHsP4ZE
T8FnsGudgrWn5OKrudKLn4tCIZP+NzzD9nsB1qtZJ712xLQlS/mxV/VusWzVPlurvtnpbHylsJZH
MpOtnUS9kVHPBeGc5JJuzEWtcUo0OJBHU/ix6DKSLkjlKHlv66eq/Vkw96rAOthXwUWL9YWk/DBZ
eyXqV/8pgKmo5O/QuDrqRSl9T++w/AVQSwMEFAAAAAgA/WB4XMMe8JnMEgAAtHcAABgAAAB4bC93
b3Jrc2hlZXRzL3NoZWV0MS54bWyd3V132ki2BuD7+RUs7gdL9V1eSWZNYehOx92Z1d1z5twSkGNO
AHkA28n8+iNAErX3LtVU+SYxu6SiXoHQozKW3v3t+3Yzeqn2h3W9ez8uJ8V4VO2W9Wq9+/p+/M8/
538149HhuNitFpt6V70f/6gO4799+Mu713r/7fBYVcdR08Ht4WmxbBqf9tWh2r9U41Nxd3g/fjwe
n25vbg7Lx2q7OEzqp2rXtDzU++3i2Dzcf705NOssVueetpsbVhTqZrtY79oebvcpfdQPD+tldVcv
n7fV7njpZF9tFscm0+Fx/XToevu+SupvtV+8Nvm78XhDvLu09P2VgvS3XS/39aF+OE6W9bYdGk1p
byzI+X2b1lH1fVkFNtN2mZJru9h/e376a9PdU7Ntvqw36+OPc1/j0XZ5+/Hrrt4vvmya17HJtVh6
IRe0+/SYTU8f3p1r/9h/eFc/HzfrXfWP/ejwvG0G9MNVm/q1eeeNu8Lv66+Px1Ph5sO7m3691bp5
bU9v0tG+eng//nt5+7lk5rTMeZH/WVevB+/n0XHx5Y9qUy2P1erc+en9+qWuv50aPzal5n1+eKxf
f9qvV/fNeJr36nH/XF2Kv9ev03rzcxOl2S/Oax/rp/vq4TitNpvTc58CnTs/jWfR/PdS9U2jw7+7
Ef7v/K4shJFaXbOcnt//uRv0/Pw6NZtlVT0snjfHZgw/V+2WEBMx7urNwP61Xh0f34/NROqCl0yO
R+1Gva9eqs3vp61ZwFqz0qnWDGJZbw7nf0fb9e4cbbv4ftlAl15LMynteLR8PhzrbftM7ZY5/ji9
N8q2m0sHrO2AXTtQEyPTO+BtB/zagZhwnt6BaDsQ1w7KCRPpHci2A+lvg4z1Vbu+8hNoHeuAgQ50
24G+dsCb1za9A9N2YPwOMta37fr2uj6bsOhrADsoi+59VKS/CqiL/q1Y+l2YMqOL7s1Yeu/GYmJz
uujejiX3RxHdIVAP3fuxFP7W5Cqji+4dWUp/EAnviJvL7n05VC2Oiw/v9vXraH/5dH1anI7I5W3Z
7539B8x4dP6PT0zzybY8rXD+IDvvlU3Doam+fCje3bycnqJdwl2WEN4SJVxi2i1x0xbucGHmFW6a
ofbjZdnjZeeepDcahsZ7WUJ5S3A03m6Jfry4MPMKYLw8e7ycjFeg8fL2uc5btpigzT/leLC4MOMD
gxXZgxVksBIN9rKE9pZQaLzdEv14cWHmFcB4ZfZ45bkn441Go/FelrDeEgaNt1uiHy8uzLwCGK/K
Hq8691QWfd+OVKakckcqM78ChqSzh6TJkEhlSip3pDIjlXlbKfvKT6TyM6l8JJVfSOUTqdz7FbBB
TPYGMeQD0aL31GWJkl03EKnckcqMVOZtxX+qEu39P/XL9JuMVD6Syi+k8olU7v0K2GQ2e5NZ8rFR
okOEuyxSXg8JU1K5I5UZqcxtYJuhA8DHdhl53R6k8olU7i3eQr+1FfrBetJP7mG2oNuI4wNtQTYS
Ld3R0oyW5l0JbCh05Pl4XajfUrT0iZbuaem3rhTYWm9ASUk+2Et8JGqXKbW3tUjpjpZmtDTvSt5u
RksfaekXWvpES/e09FtXCmyuJBOx4nz21SzE+kUup5anRfrteHFMCTakwhuyXcg/SJYaE++yECv8
hdCh9K5bqPQXQp+es9CYGPrYm3c9sQEV/dQt4L9ipPSxKwnvFet7vr5ipHRPS7/Svn6jpc+gBF/X
NDvKiby+fBfZMRVSt+tafYwxQvN2IePtEm3JhrQ/a1t5EWqdt63Nf7AVJk2DJ0gq/KRY9F0rSMpw
UkGTimhSEU0qUpKmkRUklX5ScgiQgaT49KVbyE8qo0llNKlMSZqGXZBU+UnxWU/XCpIKnFTRpCqa
VEWTqpSkaYYGSbWflByodCCpxEk1TaqjSXU0qU5JmoZjkNT4ScmRxASS4vPDbiE/qYkmNdGkJiVp
mmlBUusnxaeVXStISo6Zlia10aQ2mtQmJGVpNPWTssJPio7prmsFSfE5dLeQl7QrhZO2rQNJWZGS
NI2VIGnpJ8WndV0rSGpx0pImLaNJy2jSMiVp2iwZSMqAHPCUXtfsR+V44qlbyI/KolFZNCpLiZqP
JAaRhJXEAkriWEmMKolFlcSiSmIpSmL5SmJASfi817EAkzhmEqNMYlEmsSiTWAqTWD6TGGASOVVm
ASdxMs1LncSiTmJRJ7EUJ7F8JzHgJHyS7lgAShxDiVEosSiUWBRKLAVKLB9KDECJnNKzgJQ4lhKj
UmJRKbGolFiKlFi+lBiQEjnpZgEqcUwlRqnEolRiUSqxFCqxfCoxQCU8deBYwEocW4lRK7GolVjU
SizFSjzfShxYCU+AOB7AEsdY4hRLPIolHsUST8ESz8cSB1jC0ziOB7TEsZY41RKPaolHtcRTtMTz
tcSBlvBklOMBLQnyazqqJR7VEo9qiadoiedriQMt4ekixwNaElhLnGqJR7XEo1riKVri+VricE4J
a4kHtCSwljjVEo9qiUe1xFO0xPO1xIGW8ISR4wEtCawlTrXEgZbwWcSMR7nEU7jE87nEAZfwlJHj
AS4JzCVOucRVPGvUSzzFSzzfSxx4CU8aOR7wksBe4tRLXMezRsHEU8DE88HEAZjwtJHjATAJDCZO
wcRNPCsU06RQOG67Pm2HgfPZxAGb8OyR4wE2CcwmTtnEbTwwdNMEb+h5tz5th18uyceTAHjCk0hO
BPAkMJ4ExZMoooEF1NNEog+Debc+bYeB8wklAKHwXJITAUIJTChBCSXKeGBoqIktceASBPbaYeB8
SAkAKTyj5EQAUhJDSlBICRYPDCRVTvAp47xbn7bDwPmcEoBTeF7JiQCnJOaUoJwSPB6Yw8CGvMIc
BDZDr3A+qgRAFZ5dciKAKolRJSiqhIgHBqpiE7wrzbv1aTsMnE8rAX9fh2klArSSmFaC0krEaSUk
DKwNDixBYK8dBs73lQC+wjNNTgR8hT9Sp4L6SsR9JYCv+AQfHebd+rQdBs5HlgDIwh8eTgSQJTGy
BEWWiCNLaBhYkw8tDQLroQ+tfGkJIC086+REQFqSfMuTSkvEpSWAtMQEn3vOhS8t0A4D50tLAGnh
uScnAtKSWFqCSkvEpSWAtOSkYDiwLy3QDr8Wmy8tCaSFZ6CcDEhLku/FUmnJuLRkAQNjy8xlAQKL
gU9pmS8tCaSF56GcDEhLYmlJKi0Zl5YsYWCLX2FZgsB26BXOl5YE0sKzUU4GpKWwtCSVloxLSwJp
qQk+HM6lLy3QDgPnS0sCaeE5KScD0lJYWpJKS8alJTkMrDQOzEFgrx0GzpeWBNLCM1NOBqSlsLQk
lZaMS0sKGJi+pQUIPPiWzpeWBNLC81NOBqSlsLQklZaMS0sCaelJQV5hX1qgHQbOl5aEX5DC0pIB
aSksLUmlJePSkgoGJsdhqUDgoeOwzJeWBNLCM1VOBqSlsLQklZaMS0tqGBifpc2lBoH5UOB8aUkg
LTxd5WRAWgpLS1Jpybi0pEGBsaWlgYEHLC3zpSWBtPB0lZMBaeGP1Kmk0pJxaUkLA+NZo7m0ILAY
CKzypaWAtPATOxWQlsLSUlRaKi4tVcDA+HxkrgoQWA6cPKh8aSkgLUw8pwLSUlhaikpLxaWlShgY
y1KVIO8ALFW+sxRwFp6scirgLI2dpaizVNxZiv2XuCwpbr6yFFAWnqpyKqAsjZWlqLJUXFmKw7hk
PktxkHdoPkvlK0sBZeGpKqcCytJYWYoqS8WVpQQMjM2hBMg7QA6VbywFjIUnqpwKGEtjYylqLBU3
lkLGIlPwChpraApe5RtLAWPhJ3YqYCyNjaWosVTcWAoZC8+izRU0Fh8KnG8sBb+Ljo2lAsbCU09T
RY2l4sZSyFj4u4lzBY1VDpw2qHxjKWAsPFHlVMBYGhtLUWOpuLEUMJaaGPxrNOUbC7TDwPnGUsBY
eKLKqYCxNDaWosZScWMpCwPj+aK5siCwHDCWzjeWBsbCT+x0wFh4vnyqqbF03Fi6gIHJL0p1AQIP
/aJU5xtLA2PhiSqnA8bS2FiaGkvHjaXxbBZWh/aRBdph4HxlaaAsPFHldEBZBitLU2XpuLI0g4Gx
U+eagcBqKHC+szRwFp6ocjrgLIyiqabO0nFnaQ4D45nwueYgMB+Ah853lgbOwhNVTgecZbCzNHWW
jjtLA2fhkyTtK2vgDEnnG0sDY+EpKqcDxjLYWJoaS8eNpYGxBP181r6xxPDnc76xNDAWnqJyOmAs
g42lqbF03FhawcD42+RzrUDgcugVzjeWBsbCU1ROB4xlsLE0NZaOG0sDYzH6+yTtG4sN/j5J5xtL
wz/7w8bSAWNhEE01NZaOG0sDY5UTDPO59o0F2mHgfGNpYCw8ReV0wFgGG0tTY+m4sbSFgfGk/1xb
EFgMvMIm31gGGAtPUTkTMJbBxjLUWCZuLFPAwPjvfOamAIHLgd8nmXxjGWAsfOh3JmAsg41lqLFM
3FgGfTcLfw7OjW8s0A4D5xvLAGPhSSpnAsay2FiGGsvEjWXgt9wnGLNzw0BgOWAsk28sA4yFp6mc
CRgLf1VuaqixTNxYBn7XfYK/XjE3HATmA6cNJt9YBhgLT1M5EzAW/vXe1FBjmbixDPzGOz0OGwEC
Dx2HTb60DJAWnqhyJiAti6VlqLRMXFoGfu2dHoeNBIGHjsMmX1oGSAsfD50JSMtiaRkqLROXlol+
9934zBr4OrjJN5YBxsJTVM4EjGWxsQw1lokby0S/+m50QtR8XRmgKzw55UxAVxbrylBdmbiuTPRv
BY1JiJrvKgOvqoBdZQKusthVhrrKAFeRpNE/FTT2vye1+aCyAFR4PsrZAKgsBpWloLJFLKmN/qWg
LRKS5kvKAknhiShnA5LCU0tTSyVly2jS6B8K2jIhaT6hLCAUnoFyNkCossCGstRQlkWjRv9Q0LKE
qPl4sgBPeO7J2QCeygLryVI9WR6NGv1DQcsTouazyQI24VknZwNsKvExf2qpm6yIRo3+oaAVCVHz
wWQBmPCsk7MBMJUFFpOlYrIyGjX6d4JWJkTNp5IFVMLnWc4GqFQW2EqWWsmqaNQolWwClWw+lSyg
Ep5pcjZApbLAVrLUSlZHo0apZBOoZPOpZAGV8ByTswEqlfhP+qaWWsmaaNQolWwClWw+lSygEp5d
cjZApRJ/eXFqqZVs1Eo2aiWbYKXmiJed9bLONSzWUt8O02Iv9Yv5V9wqomLqmoeuuVUkmKk58L0h
MVATnlpyfTtMjN3ULwYSR+XUNQ8mTrBTc/x7Q2KgJzy35Pp2kBiftkz7xUDiKKC65sHECYRqDoNv
SAwQhSeXXN8OE5NLeBaUUX1tKHH8Kp5FgqSao+EbEgNL4dkl17fDxORSngXVVF8bShy/mmeRAKrm
oPiGxIBUeHrJ9e0wMbmkZ0FR1deGEkdZ1a8dT/yGy3oWQFZ4fsn17TAxubRnQW3V14YSx6/uWSTw
qjlEviExABaeZnJ9O0xMLvFZUGL1taHEUWT1a8cTv+Eyn0UrnjI42+T6dgYSk0t9dotxP3FbEwOJ
22Y5kDgBW81hMimxd8Vxd1nnfEESev3y7Mu9u7K9HHqwuzQtwO7K4e6y76HiyvYq1cHusm9x4sru
mtGh7rJvQuLK7sLMoe7SPqFhd3K4u+xbeLiyu8RwqLu0zxbYnR7uLvvmFa7sLpYb6u4Ne0U5vFck
XmkWdMeG94rEy7nC7ob3isRrpsLuhveKxOuSwu6G94rEa3/C7ob3isTra8LuhveKxGtYwu6G94rE
60TC7ob3isRrMcLu6F5x4908q735YH1s7/R3fvx+/LDYHKrxqP7yf0390D8+LKvdYr+ur5XLvRdP
dwYktXrzvN3h6u/167W03h2qPVnwUg0s+POPp2q/We++XRtW1aY6VriHSxX0cLmX4X29/Fat4GgP
9f4aePF8rOfrzbHa96Wn9Ut9/PN000jc2z93G9pfs5W31f7r+WaJB+/nyz0dXXk7K8EybZ3dzlio
zm9nPFQXtzMRqsvbmQzU/65uZzq0vLmdmVDd3s5sqF4Wt7PLnY5Iyyna5S5E/gZ42q93x89P5/uU
jr5eb0XZbsm+8sfpbXe+3VvTw9Pia/XrYv+1ed1Hm+qhaSlON/vbX97n55+P9dP5p2Yv+FIfm72g
e/R4vqfl6ZFsDgmNRhhXjJ1OCR7q+hhsaZ+xGcLz0+hp0bzL/lj/pzrfDqjer6vd8Xyb1ffjp+at
sl+sj6cdYXG+S+JpjuRhffyz9vbB8+P2/nXNw1PPn/fnIa3q192fj9Xu80u1Pz3pZajz87BGq/XD
Q7VvnuzzajV7qXbXd3PXMF/vD0dvX2yG8K/18fGuXnY3ylts1l93p1q78dr6h3f1anW512fz6ng/
Nz9env1S7n6umqfvF/cfnH7uV/AfPJyG1q8CHp0f9CvBR/4GaB72N9z98P9QSwMEFAAAAAgA/WB4
XM1LUiJ4AAAAjQAAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc02MMQ4C
IRAAe19BtvdAC2PMcdf5AKMP2HArEGEhLDH6eyktJ5OZef3kpN7UJBa2cJgMKGJXtsjewuN+3Z9B
SUfeMBUmC18SWJfdfKOEfTQSYhU1JiwWQu/1orW4QBllKpV4mGdpGfvA5nVF90JP+mjMSbf/B+jl
B1BLAQIDBhQAAAAIAP1geFz81L63YwEAAG0FAAATAAAAAAAAAAAAIAAAAAAAAABbQ29udGVudF9U
eXBlc10ueG1sUEsBAgMGFAAAAAgA/WB4XKRSuaX2AAAA3gIAAAsAAAAAAAAAAAAgAAAAlAEAAF9y
ZWxzLy5yZWxzUEsBAgMGFAAAAAgA/WB4XIQksVbpAAAAuQIAABoAAAAAAAAAAAAgAAAAswIAAHhs
L19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAgMGFAAAAAgA/WB4XIaTPmqBAQAAJQMAABAAAAAA
AAAAAAAgAAAA1AMAAGRvY1Byb3BzL2FwcC54bWxQSwECAwYUAAAACAD9YHhc5gxRVm0BAADjAgAA
EQAAAAAAAAAAACAAAACDBQAAZG9jUHJvcHMvY29yZS54bWxQSwECAwYUAAAACAD9YHhcBdGXcGgB
AADsAwAAEwAAAAAAAAAAACAAAAAfBwAAZG9jUHJvcHMvY3VzdG9tLnhtbFBLAQIDBhQAAAAIAP1g
eFxzkXtZuwUAAKYbAAATAAAAAAAAAAAAIAAAALgIAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAgMG
FAAAAAgA/WB4XALEMHQhAwAADQ0AABQAAAAAAAAAAAAgAAAApA4AAHhsL3NoYXJlZFN0cmluZ3Mu
eG1sUEsBAgMGFAAAAAgA/WB4XIYDIeXtAwAAsTkAAA0AAAAAAAAAAAAgAAAA9xEAAHhsL3N0eWxl
cy54bWxQSwECAwYUAAAACAD9YHhc7hj9ea4BAAD0AgAADwAAAAAAAAAAACAAAAAPFgAAeGwvd29y
a2Jvb2sueG1sUEsBAgMGFAAAAAgA/WB4XMMe8JnMEgAAtHcAABgAAAAAAAAAAAAgAAAA6hcAAHhs
L3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQIDBhQAAAAIAP1geFzNS1IieAAAAI0AAAAjAAAAAAAA
AAAAIAAAAOwqAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc1BLBQYAAAAADAAM
ABIDAAClKwAAAAA=
""".strip()


TELANGANA_PLANTS = {"BHUPALPALLY", "KASIPET", "KOTHAGUDEM", "KOTHAGUDAM"}
SIRMOUR_PLANT = "SIRMOUR"
OSEPL_PLANT = "OSEPL"
OSEPL_CAPACITY_MW = 20
TELANGANA_PLANT_CAPACITY_MW = {
    "BHUPALPALLY": 10,
    "KASIPET": 15,
    "KOTHAGUDEM": 37,
    "KOTHAGUDAM": 37,
}
TELANGANA_PLANT_TEMPLATE_META = {
    "BHUPALPALLY": {
        "generator": "Singareni",
        "plant_name": "Singareni Collieries Company Limited-Chelpur",
        "capacity_mw": 10,
        "contract_type": "Mtoa",
        "approval_no": "TSTRANSCO/21/2023-24",
        "to_utility": "SCCL(BPL-003, BPL-006, BPL-028)",
    },
    "KASIPET": {
        "generator": "Singareni",
        "plant_name": "Singareni Collieries Company Limited-Kasipet Mines",
        "capacity_mw": 15,
        "contract_type": "Lta",
        "approval_no": "TSTRANSCO/20/2023-24",
        "to_utility": "SCCL(BPL-003, BPL-004, BPL-065)",
    },
    "KOTHAGUDEM": {
        "generator": "Singareni",
        "plant_name": "Singareni Collieries Company Limited-Sitarampatnam",
        "capacity_mw": 37,
        "contract_type": "Lta",
        "approval_no": "TGTRANSCO/17/2024-25",
        "to_utility": "General",
    },
    "KOTHAGUDAM": {
        "generator": "Singareni",
        "plant_name": "Singareni Collieries Company Limited-Sitarampatnam",
        "capacity_mw": 37,
        "contract_type": "Lta",
        "approval_no": "TGTRANSCO/17/2024-25",
        "to_utility": "General",
    },
}


@dataclass(frozen=True)
class ConvertedAttachment:
    filename: str
    content_bytes: bytes


def _parse_csv_rows(text: str) -> List[List[str]]:
    reader = csv.reader(io.StringIO(text))
    return [[str(cell) for cell in row] for row in reader]


def _safe_get(rows: Sequence[Sequence[Any]], r: int, c: int) -> str:
    try:
        return str(rows[r][c])
    except Exception:
        return ""


def _is_numeric_cell(value: Any) -> bool:
    raw = str(value if value is not None else "").strip()
    if raw == "":
        return False
    try:
        float(raw)
        return True
    except Exception:
        return False


def _to_float_or_none(value: Any) -> Optional[float]:
    raw = str(value if value is not None else "").strip().replace(",", "")
    if raw == "":
        return None
    try:
        num = float(raw)
    except Exception:
        return None
    return num if math.isfinite(num) else None


def _extract_iso_date(value: Any) -> str:
    match = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", str(value or ""))
    return match.group(1) if match else ""


def _resolve_telangana_schedule_date(rows: Sequence[Sequence[Any]], report_date: str = "") -> str:
    for row in rows:
        for cell in row:
            date_text = _extract_iso_date(cell)
            if date_text:
                return date_text
    return _extract_iso_date(report_date)


def _telangana_block_start_timestamp(block: int, schedule_date: str = "") -> str:
    start_minutes = max(0, int(block) - 1) * 15
    hour, minute = divmod(start_minutes, 60)
    time_text = f"{hour % 24:02d}:{minute:02d}:00"
    return f"{schedule_date} {time_text}" if schedule_date else time_text


def _format_telangana_template_date(schedule_date: str = "") -> str:
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", str(schedule_date or "").strip())
    if not match:
        return schedule_date
    year, month, day = match.groups()
    return f"{day}-{month}-{year}"


def _telangana_day_ahead_revision(template_id: str = "", schedule_type: str = "") -> Optional[int]:
    selector = f"{template_id} {schedule_type}".lower()
    if re.search(r"(?:^|[_\s-])da0(?:$|[_\s-])", selector):
        return 0
    if re.search(r"(?:^|[_\s-])da1(?:$|[_\s-])", selector):
        return 1
    return None


def _is_telangana_day_ahead_manual_source(
    *,
    plant_code: str = "",
    template_id: str = "",
    schedule_type: str = "",
    source_key: str = "",
    file_name: str = "",
) -> bool:
    plant = str(plant_code or "").strip().upper()
    if plant not in TELANGANA_PLANTS:
        return False
    if _telangana_day_ahead_revision(template_id, schedule_type) not in {0, 1}:
        return False
    source_text = f"{source_key} {file_name}".lower()
    return "manual-edits/" in source_text or "edited_schedule" in source_text


def _telangana_schedule_values_by_block(rows: Sequence[Sequence[Any]]) -> Dict[int, Any]:
    def norm(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    header_idx = -1
    header: Sequence[Any] = []
    for idx, row in enumerate(rows[:80]):
        normalized = [norm(cell) for cell in row]
        if "block" in normalized and any(
            token in normalized
            for token in ("mw", "stationschedule", "scheduledmw", "schedule", "forecastmw", "forecast")
        ):
            header_idx = idx
            header = row
            break

    normalized_header = [norm(cell) for cell in header]
    if header_idx >= 0:
        block_col = normalized_header.index("block") if "block" in normalized_header else 0
        schedule_col = -1
        for candidate in ("mw", "stationschedule", "scheduledmw", "schedule", "forecastmw", "forecast"):
            if candidate in normalized_header:
                schedule_col = normalized_header.index(candidate)
                break
        if schedule_col < 0:
            schedule_col = 1 if len(header) > 1 else 0
        data_rows = rows[header_idx + 1:]
    else:
        block_col = 0
        schedule_col = 1
        data_rows = rows

    values: Dict[int, Any] = {}
    for row in data_rows:
        raw_block = row[block_col] if block_col < len(row) else ""
        block_num = _to_float_or_none(raw_block)
        if block_num is None:
            continue
        block = int(block_num)
        if block < 1 or block > 96:
            continue
        raw_value = row[schedule_col] if schedule_col < len(row) else ""
        if _to_float_or_none(raw_value) is None:
            continue
        values[block] = raw_value
    return values


def _load_telangana_template_workbook() -> "openpyxl.workbook.workbook.Workbook":
    from openpyxl import load_workbook  # type: ignore

    # The B64 blob is filled in at build time by the repo patch (see apply_patch below).
    tpl = base64.b64decode(re.sub(r"\s+", "", _TELANGANA_TEMPLATE_XLSX_B64).encode("ascii"))
    return load_workbook(io.BytesIO(tpl))


def _write_workbook_bytes(wb: "openpyxl.workbook.workbook.Workbook") -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _derive_telangana_avc_by_block(rows: Sequence[Sequence[Any]], plant_code: str) -> dict[int, float]:
    capacity = TELANGANA_PLANT_CAPACITY_MW.get(str(plant_code or "").strip().upper())
    if not capacity:
        return {}

    def norm(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    header_idx = -1
    header: Sequence[Any] = []
    for idx, row in enumerate(rows[:80]):
        normalized = [norm(cell) for cell in row]
        joined = " ".join(normalized)
        if "block" in joined and ("stationschedule" in joined or "forecast" in joined or "schedule" in joined):
            header_idx = idx
            header = row
            break

    if header_idx < 0:
        header_idx = 11
        header = rows[header_idx] if header_idx < len(rows) else []

    normalized_header = [norm(cell) for cell in header]

    def find_col(candidates: Sequence[str], fallback: int) -> int:
        for candidate in candidates:
            candidate_norm = norm(candidate)
            for col_idx, value in enumerate(normalized_header):
                if value == candidate_norm or (candidate_norm and candidate_norm in value):
                    return col_idx
        return fallback

    block_col = find_col(["Block"], 0)
    schedule_col = find_col(["Station Schedule", "Scheduled MW", "Schedule", "Forecast(MW)", "Forecast"], 4)
    active_blocks: List[int] = []

    for row in rows[header_idx + 1:]:
        raw_block = row[block_col] if block_col < len(row) else ""
        block_num = _to_float_or_none(raw_block)
        if block_num is None:
            continue
        block = int(block_num)
        if block < 1 or block > 96:
            continue
        schedule_value = row[schedule_col] if schedule_col < len(row) else ""
        schedule_num = _to_float_or_none(schedule_value)
        if schedule_num is not None and schedule_num > 0:
            active_blocks.append(block)

    if not active_blocks:
        return {}

    first = min(active_blocks)
    last = max(active_blocks)
    return {block: float(capacity if first <= block <= last else 0) for block in range(1, 97)}


def convert_telangana_csv_to_sldc_xlsx_bytes(
    csv_text: str,
    *,
    sheet_name: str = "SLDC Template",
    plant_code: str = "",
    template_id: str = "",
    schedule_type: str = "",
    fill_blank_avc: bool = False,
    report_date: str = "",
    source_key: str = "",
    file_name: str = "",
) -> bytes:
    """
    Convert schedule CSV (as returned by `/email-scheduler/resolve-s3-schedule-attachment`)
    into the Telangana SLDC XLSX format.

    This mirrors the frontend behavior used by the Email Scheduler screen.
    """
    from openpyxl.styles import Alignment, Font  # type: ignore

    rows = _parse_csv_rows(csv_text)
    day_ahead_revision = _telangana_day_ahead_revision(template_id, schedule_type)
    is_telangana_day_ahead = (
        str(plant_code or "").strip().upper() in TELANGANA_PLANTS
        and day_ahead_revision in {0, 1}
    )
    has_resolved_source_key = bool(str(source_key or "").strip())
    forced_report_date = _extract_iso_date(report_date) if is_telangana_day_ahead and has_resolved_source_key else ""
    schedule_date = forced_report_date or _resolve_telangana_schedule_date(rows, report_date=report_date)
    is_telangana_da1 = is_telangana_day_ahead and day_ahead_revision == 1
    is_day_ahead_manual_source = _is_telangana_day_ahead_manual_source(
        plant_code=plant_code,
        template_id=template_id,
        schedule_type=schedule_type,
        source_key=source_key,
        file_name=file_name,
    )
    manual_values_by_block = _telangana_schedule_values_by_block(rows) if is_day_ahead_manual_source else {}
    derived_avc_by_block = _derive_telangana_avc_by_block(rows, plant_code) if fill_blank_avc else {}
    fallback_avc_capacity = TELANGANA_PLANT_CAPACITY_MW.get(str(plant_code or "").strip().upper()) if fill_blank_avc else None
    wb = _load_telangana_template_workbook()
    ws = wb.worksheets[0]
    ws.title = sheet_name

    def writable_cell(r: int, c: int) -> Any:
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= r <= merged_range.max_row and merged_range.min_col <= c <= merged_range.max_col:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        return ws.cell(row=r, column=c)

    def set_text(r: int, c: int, v: str) -> None:
        writable_cell(r, c).value = v if v != "" else ""

    def set_number(r: int, c: int, v: str, *, blank_if_empty: bool = False) -> None:
        cell = writable_cell(r, c)
        raw = "" if v is None else str(v).strip()
        if raw == "" and blank_if_empty:
            cell.value = None
            return
        try:
            cell.value = float(raw)
        except Exception:
            cell.value = None if blank_if_empty else 0

    plant_key = str(plant_code or "").strip().upper()
    plant_meta = TELANGANA_PLANT_TEMPLATE_META.get(plant_key, {})
    set_text(1, 2, str(plant_meta.get("generator") or _safe_get(rows, 0, 1)))
    set_text(2, 2, str(plant_meta.get("plant_name") or _safe_get(rows, 1, 1)))
    set_number(3, 2, str(plant_meta.get("capacity_mw") or _safe_get(rows, 2, 1)), blank_if_empty=False)
    set_text(4, 2, _format_telangana_template_date(schedule_date) or _safe_get(rows, 3, 1))
    set_text(5, 2, "dayahead")
    if day_ahead_revision is not None:
        set_number(6, 2, str(day_ahead_revision), blank_if_empty=False)
    set_text(8, 6, str(plant_meta.get("contract_type") or _safe_get(rows, 7, 5)))
    set_text(9, 6, str(plant_meta.get("approval_no") or _safe_get(rows, 8, 5)))
    set_text(10, 6, str(plant_meta.get("to_utility") or _safe_get(rows, 9, 5)))
    set_number(12, 6, str(plant_meta.get("capacity_mw") or _safe_get(rows, 11, 5)), blank_if_empty=False)

    source_rows_by_block = {}
    for candidate in rows[12:]:
        block_num = _to_float_or_none(candidate[0] if len(candidate) > 0 else "")
        if block_num is None:
            continue
        block_key = int(block_num)
        if 1 <= block_key <= 96 and block_key not in source_rows_by_block:
            source_rows_by_block[block_key] = candidate
    has_source_block_map = bool(source_rows_by_block)

    # Fill 96 blocks into A13..F108 when present; otherwise leave template blanks.
    for i in range(96):
        block_key = i + 1
        if has_source_block_map:
            src = source_rows_by_block.get(block_key, [])
        else:
            src = rows[12 + i] if (12 + i) < len(rows) else []
        excel_row = 13 + i
        # Columns: A block, B time, C forecast, D AvC, E station schedule, F capacity/helper
        set_number(excel_row, 1, str(block_key))
        set_text(excel_row, 2, _telangana_block_start_timestamp(block_key, schedule_date))
        source_avc = src[3] if len(src) > 3 else ""
        source_avc_num = _to_float_or_none(source_avc)
        if (source_avc_num is None or abs(source_avc_num) < 1e-9) and derived_avc_by_block:
            source_avc = derived_avc_by_block.get(block_key, 0)
        elif fallback_avc_capacity is not None and (source_avc_num is None or abs(source_avc_num) < 1e-9):
            source_avc = fallback_avc_capacity
        set_number(excel_row, 4, source_avc, blank_if_empty=True)
        station_schedule = src[4] if len(src) > 4 else ""
        if fill_blank_avc and _to_float_or_none(station_schedule) is None:
            station_schedule = 0
        if is_day_ahead_manual_source:
            station_schedule = manual_values_by_block.get(block_key, station_schedule)
        forecast_value = station_schedule if is_day_ahead_manual_source else ("" if is_telangana_da1 else (src[2] if len(src) > 2 else ""))
        set_number(excel_row, 3, forecast_value, blank_if_empty=True)
        set_number(excel_row, 5, station_schedule, blank_if_empty=True)
        helper_value = station_schedule if fill_blank_avc else (src[5] if len(src) > 5 else "")
        set_number(excel_row, 6, helper_value, blank_if_empty=True)

    # Normalize some basic visual alignment similar to the UI (kept light).
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for r in range(1, 109):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align
            if r == 12 and c != 6:
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = Font(bold=False, size=11)

    return _write_workbook_bytes(wb)


def _format_sirmour_number(value: Any) -> Any:
    raw = str(value if value is not None else "").strip()
    if raw == "":
        return 0
    try:
        num = float(raw)
    except Exception:
        return raw
    if abs(num - int(num)) < 1e-9:
        return int(num)
    return num


def _format_osepl_number(value: Any) -> str:
    num = _to_float_or_none(value)
    if num is None:
        return "0"
    if abs(num - int(num)) < 1e-9:
        return str(int(num))
    return f"{num:.6f}".rstrip("0").rstrip(".")


def _osepl_day_ahead_revision(template_id: str = "", schedule_type: str = "") -> str:
    selector = f"{template_id} {schedule_type}".lower()
    if re.search(r"(?:^|[_\s-])da0(?:$|[_\s-])", selector):
        return "DA"
    if re.search(r"(?:^|[_\s-])da1(?:$|[_\s-])", selector):
        return "DA"
    return "DA"


def _extract_source_schedule_by_block(rows: Sequence[Sequence[Any]]) -> dict[int, float]:
    header_idx = -1
    best_score = -1
    for idx, row in enumerate(rows[:100]):
        normalized = [_normalize_header(cell) for cell in row]
        joined = " ".join(normalized)
        score = 0
        if "block" in joined:
            score += 4
        if "stationschedule" in joined or "schedule" in joined or "forecast" in joined or "declaredforecast" in joined:
            score += 4
        if score > best_score:
            best_idx = idx
            best_score = score
            header_idx = best_idx
    header = rows[header_idx] if header_idx >= 0 and header_idx < len(rows) else []
    block_col = _column_index(header, ["Block"])
    schedule_col = _column_index(header, ["Station Schedule", "Schedule", "Declared Forecast", "Forecast(MW)", "Forecast", "Scheduled MW", "Scheduled"])
    if block_col < 0:
        block_col = 0
    if schedule_col < 0:
        schedule_col = 4 if len(header) > 4 else 1

    values: dict[int, float] = {}
    data_start = header_idx + 1 if header_idx >= 0 else 0
    for row in rows[data_start:]:
        raw_block = row[block_col] if block_col < len(row) else ""
        block_num = _to_float_or_none(raw_block)
        if block_num is None:
            continue
        block = int(block_num)
        if block < 1 or block > 96:
            continue
        raw_value = row[schedule_col] if schedule_col < len(row) else ""
        value = _to_float_or_none(raw_value)
        values[block] = float(value or 0)
    return values


def convert_osepl_day_ahead_csv_bytes(
    csv_text: str,
    *,
    template_id: str = "",
    schedule_type: str = "",
    report_date: str = "",
) -> bytes:
    rows = _parse_csv_rows(csv_text)
    schedule_date = _resolve_telangana_schedule_date(rows, report_date=report_date)
    values_by_block = _extract_source_schedule_by_block(rows)
    active_blocks = [block for block, value in values_by_block.items() if math.isfinite(value) and value > 0]
    first_active = min(active_blocks) if active_blocks else None
    last_active = max(active_blocks) if active_blocks else None
    revision = _osepl_day_ahead_revision(template_id, schedule_type)

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow([f"Schedule Template for MH_VEDANJAY and revision {revision}"])
    writer.writerow(["", "Scheduling entity", "MH_VEDANJAY"])
    writer.writerow(["", "Date", schedule_date])
    writer.writerow(["", "Revision No", revision])
    writer.writerow([])
    writer.writerow(["POS Name", "Naldurg Inter 132kV", "Naldurg Inter 132kV", "Naldurg Inter 132kV"])
    writer.writerow(["Down Stream Name", "", "", "Naldurg Inter 132kV"])
    writer.writerow(["Energy Type", "", "", "SOLAR"])
    writer.writerow(["Contract ID", "", "", "CONTRACT00192"])
    writer.writerow(["Contract Type", "", "", "LTA"])
    writer.writerow(["Exchange Type", "", "", "NA"])
    writer.writerow(["Transaction Type", "INTER", "INTER", "INTER"])
    writer.writerow(["RE Generator Name", "", "", "Naldurg Inter 132kV"])
    writer.writerow(["Path", "", "", "WR-WR"])
    writer.writerow(["Buyer Name", "", "", "SOLAR_CSEB"])
    writer.writerow(["STU Name", "", "", "Naldurg 132kV"])
    writer.writerow(["Approval Number", "", "", "L_WR_2014_03"])
    writer.writerow(["Capacity", OSEPL_CAPACITY_MW, OSEPL_CAPACITY_MW, OSEPL_CAPACITY_MW])
    writer.writerow(["Block", "Declared Forecast", "Inter Avc", "Schedule"])
    for block in range(1, 97):
        schedule_value = float(values_by_block.get(block, 0) or 0)
        inter_avc = OSEPL_CAPACITY_MW if first_active is not None and last_active is not None and first_active <= block <= last_active else 0
        writer.writerow([
            block,
            _format_osepl_number(schedule_value),
            _format_osepl_number(inter_avc),
            _format_osepl_number(schedule_value),
        ])
    return output.getvalue().encode("utf-8")


def _sirmour_time_interval(block: int) -> str:
    start_minutes = (block - 1) * 15
    end_minutes = block * 15
    start_h, start_m = divmod(start_minutes, 60)
    end_h, end_m = divmod(end_minutes, 60)
    end_h %= 24
    return f"{start_h:02d}:{start_m:02d}-{end_h:02d}:{end_m:02d}"


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _find_sirmour_table_header(rows: Sequence[Sequence[Any]]) -> int:
    best_idx = -1
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        normalized = [_normalize_header(cell) for cell in row]
        joined = " ".join(normalized)
        score = 0
        if "block" in joined:
            score += 4
        if "time" in joined or "blockinterval" in joined:
            score += 3
        if "forecast" in joined or "scheduledmw" in joined or "scheduled" in joined:
            score += 4
        if "availability" in joined:
            score += 3
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx if best_score >= 4 else -1


def _column_index(headers: Sequence[Any], candidates: Sequence[str]) -> int:
    normalized = [_normalize_header(h) for h in headers]
    for candidate in candidates:
        candidate_norm = _normalize_header(candidate)
        for idx, header in enumerate(normalized):
            if header == candidate_norm:
                return idx
    for candidate in candidates:
        candidate_norm = _normalize_header(candidate)
        for idx, header in enumerate(normalized):
            if candidate_norm and candidate_norm in header:
                return idx
    return -1


def _sirmour_meta_value(rows: Sequence[Sequence[Any]], label: str, fallback: str) -> str:
    label_norm = _normalize_header(label)
    for row in rows[:20]:
        if not row:
            continue
        if _normalize_header(row[0]) == label_norm:
            value = str(row[1] if len(row) > 1 else "").strip()
            return value or fallback
    return fallback


def _build_sirmour_sldc_rows(csv_text: str, *, report_date: str = "") -> List[List[Any]]:
    rows = _parse_csv_rows(csv_text)
    report_date = str(report_date or "").strip()
    template_date = _sirmour_meta_value(rows, "DATE:", report_date)
    revision = _sirmour_meta_value(rows, "REVISION:", "3")

    table_header = _find_sirmour_table_header(rows)
    first_header = rows[table_header] if table_header >= 0 else []
    second_header = rows[table_header + 1] if table_header >= 0 and (table_header + 1) < len(rows) else []
    merged_headers = [
        " ".join(str(part or "").strip() for part in parts if str(part or "").strip())
        for parts in zip(
            list(first_header) + [""] * max(0, len(second_header) - len(first_header)),
            list(second_header) + [""] * max(0, len(first_header) - len(second_header)),
        )
    ]
    headers = merged_headers if any("forecast" in _normalize_header(h) or "availability" in _normalize_header(h) for h in merged_headers) else first_header

    block_col = _column_index(headers, ["Block"])
    time_col = _column_index(headers, ["Block Interval", "Time"])
    availability_col = _column_index(headers, ["Availability"])
    forecast_col = _column_index(headers, ["Forecast", "Scheduled MW", "Scheduled M", "Schedule", "Scheduled"])

    values_by_block: dict[int, Tuple[str, Any, Any]] = {}
    data_start = table_header + 1 if table_header >= 0 else 0
    if second_header and any("forecast" in _normalize_header(h) or "availability" in _normalize_header(h) for h in second_header):
        data_start = table_header + 2

    for row in rows[data_start:]:
        if not row:
            continue
        raw_block = row[block_col] if 0 <= block_col < len(row) else row[0]
        try:
            block = int(float(str(raw_block).strip()))
        except Exception:
            continue
        if block < 1 or block > 96:
            continue
        time_value = str(row[time_col] if 0 <= time_col < len(row) else "").strip() or _sirmour_time_interval(block)
        forecast_value = row[forecast_col] if 0 <= forecast_col < len(row) else 0
        availability_value = row[availability_col] if 0 <= availability_col < len(row) else None
        if availability_value is None or str(availability_value).strip() == "":
            try:
                availability_value = 5.1 if float(str(forecast_value or "0").strip()) > 0 else 0
            except Exception:
                availability_value = 0
        values_by_block[block] = (time_value, _format_sirmour_number(availability_value), _format_sirmour_number(forecast_value))

    out_rows: List[List[Any]] = [
        ["TYPE:", "REG", "", ""],
        ["DATE:", template_date, "", ""],
        ["REVISION:", revision, "", ""],
        ["REASON:", "NA", "", ""],
        ["Block", "Block Interval", "5.1MW M/s SIRMOUR SMALL HYDRO POWER PVT LTD", ""],
        ["", "", "Availability", "Forecast"],
    ]
    for block in range(1, 97):
        time_value, availability_value, forecast_value = values_by_block.get(
            block,
            (_sirmour_time_interval(block), 0, 0),
        )
        out_rows.append([block, time_value, availability_value, forecast_value])
    return out_rows


def convert_sirmour_csv_to_gsnp_xlsx_bytes(csv_text: str, *, sheet_name: str = "SLDC Template", report_date: str = "") -> bytes:
    """
    Convert SIRMOUR schedule CSV to a simple GSNP-style XLSX, matching the frontend
    Email Scheduler behavior (numeric cells written as numbers).
    """
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Alignment, Font  # type: ignore

    rows = _build_sirmour_sldc_rows(csv_text, report_date=report_date)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    base_font = Font(size=11, bold=False)

    # Determine where the "table" starts (best-effort), so we can right-align key numeric columns.
    header_row = -1
    availability_col = -1
    forecast_col = -1
    for idx, row in enumerate(rows[:200]):
        lowered = [str(c or "").strip().lower() for c in row]
        if any("availability" in c for c in lowered) or any("forecast" in c for c in lowered):
            header_row = idx
            if lowered:
                availability_col = next((i for i, c in enumerate(lowered) if "availability" in c), -1)
                forecast_col = next((i for i, c in enumerate(lowered) if "forecast" in c), -1)
            break
    data_start = header_row + 1 if header_row >= 0 else 0

    for r_idx, row in enumerate(rows):
        is_revision_row = str(row[0] if row else "").strip().lower().find("revision") >= 0
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
            if _is_numeric_cell(val):
                try:
                    cell.value = float(str(val).strip())
                except Exception:
                    cell.value = str(val)
            else:
                cell.value = val
            cell.font = base_font
            if is_revision_row and c_idx == 1:
                cell.alignment = align_left
            elif r_idx >= data_start and (c_idx == availability_col or c_idx == forecast_col):
                cell.alignment = align_right
            else:
                cell.alignment = Alignment(vertical="center")

    return _write_workbook_bytes(wb)


def maybe_convert_for_auto_email(
    *,
    plant_code: str,
    template_id: str,
    schedule_type: str,
    file_name: str,
    file_bytes: bytes,
    report_date: str = "",
    source_key: str = "",
) -> Optional[ConvertedAttachment]:
    """
    Apply plant-wise SLDC XLSX conversion for cron-driven auto emails only.

    Returns ConvertedAttachment if conversion was applied; otherwise None.
    """
    plant = str(plant_code or "").strip().upper()
    if not plant or not file_bytes:
        return None

    # Only convert schedule CSVs; if the source is already XLSX, keep as-is.
    if str(file_name or "").lower().endswith(".xlsx"):
        return None

    csv_text = file_bytes.decode("utf-8", errors="replace")
    safe_type = str(schedule_type or "").strip().lower() or "schedule"
    sheet_name = f"{plant} {safe_type.upper()}".strip()
    out_base = f"{plant}_{safe_type}"
    if plant == OSEPL_PLANT:
        template_key = str(template_id or "").strip().lower()
        is_day_ahead = (
            "da0" in template_key
            or "da1" in template_key
            or safe_type in {"dayahead", "da0", "da1"}
        )
        if not is_day_ahead:
            return None
        content = convert_osepl_day_ahead_csv_bytes(
            csv_text,
            template_id=template_id,
            schedule_type=schedule_type,
            report_date=report_date,
        )
        return ConvertedAttachment(filename=f"{out_base}.csv", content_bytes=content)
    if plant in TELANGANA_PLANTS:
        template_key = str(template_id or "").strip().lower()
        # Cron callers can pass either the normalized type ("dayahead") or the
        # DA selector itself ("DA0"/"DA1"). Keep the AvC fallback tied to the
        # Telangana DA templates, not to one exact schedule_type spelling.
        fill_blank_avc = (
            "da0" in template_key
            or "da1" in template_key
            or safe_type in {"dayahead", "da0", "da1"}
        )
        content = convert_telangana_csv_to_sldc_xlsx_bytes(
            csv_text,
            sheet_name=sheet_name,
            plant_code=plant,
            template_id=template_id,
            schedule_type=schedule_type,
            fill_blank_avc=fill_blank_avc,
            report_date=report_date,
            source_key=source_key,
            file_name=file_name,
        )
        return ConvertedAttachment(filename=f"{out_base}.xlsx", content_bytes=content)
    if plant == SIRMOUR_PLANT:
        content = convert_sirmour_csv_to_gsnp_xlsx_bytes(csv_text, sheet_name=sheet_name, report_date=report_date)
        return ConvertedAttachment(filename=f"{out_base}.xlsx", content_bytes=content)

    return None
