import csv
import io
import os
import posixpath
import re
import socket
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Sequence, Tuple


SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xls")


@dataclass
class RemoteFile:
    name: str
    path: str
    modified_at: Optional[datetime]
    size: int = 0


class UtilityFileServiceError(RuntimeError):
    pass


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name) or default)
    except Exception:
        return default


def _sftp_config() -> Dict[str, Any]:
    return {
        "host": (os.getenv("UTILITY_SFTP_HOST") or "").strip(),
        "port": _env_int("UTILITY_SFTP_PORT", 22),
        "username": (os.getenv("UTILITY_SFTP_USERNAME") or "").strip(),
        "password": os.getenv("UTILITY_SFTP_PASSWORD") or "",
        "base_dir": (os.getenv("UTILITY_SFTP_BASE_DIR") or "/wbes").strip() or "/wbes",
    }


def _require_config(config: Dict[str, Any]) -> None:
    missing = [key for key in ("host", "username", "password") if not str(config.get(key) or "").strip()]
    if missing:
        raise UtilityFileServiceError(f"Missing SFTP config: {', '.join(missing)}")


def _connect_sftp():
    config = _sftp_config()
    _require_config(config)
    timeout = _env_int("UTILITY_SFTP_TIMEOUT_SECONDS", 10)
    try:
        import paramiko  # type: ignore
    except Exception as exc:
        raise UtilityFileServiceError("paramiko is required for SFTP folder scan") from exc

    sock = None
    transport = None
    try:
        sock = socket.create_connection((config["host"], int(config["port"])), timeout=timeout)
        sock.settimeout(timeout)
        transport = paramiko.Transport(sock)
        transport.banner_timeout = timeout
        transport.auth_timeout = timeout
        transport.connect(username=config["username"], password=config["password"])
        sftp = paramiko.SFTPClient.from_transport(transport)
        return transport, sftp, config
    except Exception as exc:
        try:
            if transport is not None:
                transport.close()
        finally:
            if sock is not None:
                try:
                    sock.close()
                except Exception:
                    pass
        raise UtilityFileServiceError(f"SFTP connection failed: {exc}") from exc


def _safe_segment(value: str) -> str:
    text = str(value or "").strip().strip("/\\")
    if not text or text in {".", ".."} or "/" in text or "\\" in text:
        raise UtilityFileServiceError("Invalid folder name")
    return text


def _join_remote(*parts: str) -> str:
    joined = posixpath.join(*[str(p or "").strip("/") for p in parts if str(p or "").strip("/")])
    return "/" + joined.strip("/")


def _remote_mtime(attr: Any) -> Optional[datetime]:
    try:
        ts = int(getattr(attr, "st_mtime", 0) or 0)
        return datetime.fromtimestamp(ts) if ts else None
    except Exception:
        return None


def _is_dir(attr: Any) -> bool:
    try:
        import stat

        return stat.S_ISDIR(int(getattr(attr, "st_mode", 0) or 0))
    except Exception:
        return False


def _list_attrs(sftp: Any, path: str) -> List[Any]:
    try:
        return list(sftp.listdir_attr(path))
    except FileNotFoundError as exc:
        raise UtilityFileServiceError(f"Folder not found: {path}") from exc


def list_utilities() -> Dict[str, Any]:
    transport, sftp, config = _connect_sftp()
    try:
        base_dir = config["base_dir"]
        utilities = []
        for attr in _list_attrs(sftp, base_dir):
            name = str(getattr(attr, "filename", "") or "").strip()
            if not name or name.startswith(".") or not _is_dir(attr):
                continue
            utilities.append(
                {
                    "name": name,
                    "path": _join_remote(base_dir, name),
                    "modified_at": _remote_mtime(attr).isoformat() if _remote_mtime(attr) else None,
                }
            )
        utilities.sort(key=lambda item: item["name"].lower())
        return {"items": utilities, "base_dir": base_dir}
    finally:
        sftp.close()
        transport.close()


def list_dates(utility: str) -> Dict[str, Any]:
    utility_name = _safe_segment(utility)
    transport, sftp, config = _connect_sftp()
    try:
        utility_path = _join_remote(config["base_dir"], utility_name)
        dates = []
        for attr in _list_attrs(sftp, utility_path):
            name = str(getattr(attr, "filename", "") or "").strip()
            if not name or name.startswith(".") or not _is_dir(attr):
                continue
            dates.append(
                {
                    "date": name,
                    "path": _join_remote(utility_path, name),
                    "modified_at": _remote_mtime(attr).isoformat() if _remote_mtime(attr) else None,
                }
            )
        dates.sort(key=lambda item: item["date"], reverse=True)
        return {"items": dates, "utility": utility_name}
    finally:
        sftp.close()
        transport.close()


def _list_supported_files(sftp: Any, folder: str) -> List[RemoteFile]:
    files: List[RemoteFile] = []
    for attr in _list_attrs(sftp, folder):
        name = str(getattr(attr, "filename", "") or "").strip()
        if not name or name.startswith(".") or _is_dir(attr):
            continue
        if not name.lower().endswith(SUPPORTED_EXTENSIONS):
            continue
        files.append(
            RemoteFile(
                name=name,
                path=_join_remote(folder, name),
                modified_at=_remote_mtime(attr),
                size=int(getattr(attr, "st_size", 0) or 0),
            )
        )
    files.sort(key=lambda item: (item.modified_at or datetime.min, item.name), reverse=True)
    return files


def _read_remote_bytes(sftp: Any, path: str) -> bytes:
    with sftp.open(path, "rb") as handle:
        return handle.read()


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _row_non_empty_count(row: Sequence[Any]) -> int:
    return sum(1 for value in row if str(value if value is not None else "").strip())


def _detect_header_row(rows: List[List[Any]]) -> Tuple[List[str], List[List[Any]]]:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:30]):
        normalized = [_normalize_header(cell) for cell in row]
        score = _row_non_empty_count(row)
        if any(token in normalized for token in ("block", "time", "timeslot", "oa_remc", "as")):
            score += 20
        if any("oaremc" in token for token in normalized):
            score += 15
        if "as" in normalized:
            score += 10
        if score > best_score:
            best_idx = idx
            best_score = score
    headers = [str(value if value is not None else "").strip() or f"Column {i + 1}" for i, value in enumerate(rows[best_idx])]
    data_rows = rows[best_idx + 1 :]
    width = len(headers)
    normalized_rows = [list(row[:width]) + [""] * max(0, width - len(row)) for row in data_rows]
    return headers, normalized_rows


def _parse_csv_rows(data: bytes) -> Tuple[List[str], List[List[Any]]]:
    text = _decode_text(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [list(row) for row in reader if any(str(cell or "").strip() for cell in row)]
    if not rows:
        return [], []
    return _detect_header_row(rows)


def _parse_xlsx_rows(data: bytes) -> Tuple[List[str], List[List[Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise UtilityFileServiceError("openpyxl is required for XLSX preview") from exc
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True) if any(str(cell if cell is not None else "").strip() for cell in row)]
    if not rows:
        return [], []
    return _detect_header_row(rows)


def _parse_file_rows(file_name: str, data: bytes) -> Tuple[List[str], List[List[Any]]]:
    lower = str(file_name or "").lower()
    if lower.endswith(".csv"):
        return _parse_csv_rows(data)
    if lower.endswith((".xlsx", ".xls")):
        return _parse_xlsx_rows(data)
    return [], []


def _to_float(value: Any) -> Optional[float]:
    text = str(value if value is not None else "").strip().replace(",", "")
    if not text or text.upper() in {"NA", "N/A", "#NA", "#N/A", "-"}:
        return None
    try:
        num = float(text)
    except Exception:
        return None
    return num if num == num else None


def _find_column(headers: List[str], candidates: List[str]) -> int:
    normalized = [_normalize_header(header) for header in headers]
    candidate_norm = [_normalize_header(item) for item in candidates]
    for needle in candidate_norm:
        for idx, token in enumerate(normalized):
            if token == needle:
                return idx
    for needle in candidate_norm:
        for idx, token in enumerate(normalized):
            if needle and needle in token:
                return idx
    return -1


def _filename_next_block(file_name: str) -> Dict[str, Any]:
    name = str(file_name or "")
    match = re.search(r"@(\d{1,2})[-_:](\d{2})[-_:](\d{2})@", name)
    if not match:
        match = re.search(r"captured[-_](?:\d{8})[-_](\d{2})(\d{2})(\d{2})", name, flags=re.IGNORECASE)
    if not match:
        matches = list(re.finditer(r"(?:^|[^\d])(\d{1,2})[-_:](\d{2})(?:[-_:](\d{2}))(?=$|[^\d])", name))
        match = matches[-1] if matches else None
    if not match:
        return {"block": None, "interval": "", "source_time": ""}
    hour = max(0, min(23, int(match.group(1))))
    minute = max(0, min(59, int(match.group(2))))
    source_time = f"{hour:02d}:{minute:02d}:{int(match.group(3) or 0):02d}"
    total_minutes = (hour * 60) + minute
    next_q = ((total_minutes // 15) + 1) * 15
    start = next_q % (24 * 60)
    end = (start + 15) % (24 * 60)
    block = (start // 15) + 1
    interval = f"{start // 60:02d}:{start % 60:02d}-{end // 60:02d}:{end % 60:02d}"
    return {"block": block, "interval": interval, "source_time": source_time}


def _row_block(row: List[Any], headers: List[str], idx: int) -> Optional[int]:
    block_idx = _find_column(headers, ["block", "block no", "block number"])
    if block_idx >= 0 and block_idx < len(row):
        val = _to_float(row[block_idx])
        if val is not None and 1 <= int(val) <= 96:
            return int(val)
    time_idx = _find_column(headers, ["time", "time slot", "timeslot", "block interval"])
    if time_idx >= 0 and time_idx < len(row):
        match = re.search(r"(\d{1,2}):(\d{2})", str(row[time_idx] or ""))
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            total = (hour * 60) + minute
            block = (total // 15) + 1
            if 1 <= block <= 96:
                return block
    fallback = idx + 1
    return fallback if 1 <= fallback <= 96 else None


def _compute_total_card(headers: List[str], rows: List[List[Any]], file_name: str) -> Dict[str, Any]:
    block_info = _filename_next_block(file_name)
    target_block = block_info.get("block")
    oa_idx = _find_column(headers, ["OA_REMC", "OA REMC", "OA-REMC", "OAREMC"])
    as_idx = _find_column(headers, ["AS"])
    target_row: Optional[List[Any]] = None
    if target_block:
        for idx, row in enumerate(rows):
            if _row_block(row, headers, idx) == target_block:
                target_row = row
                break
    if target_row is None and rows:
        target_row = rows[0]
    oa = _to_float(target_row[oa_idx]) if target_row is not None and oa_idx >= 0 and oa_idx < len(target_row) else None
    as_value = _to_float(target_row[as_idx]) if target_row is not None and as_idx >= 0 and as_idx < len(target_row) else None
    total = (oa or 0.0) + (as_value or 0.0)
    return {
        **block_info,
        "oa_remc": oa,
        "as": as_value,
        "total": total,
        "oa_remc_column": headers[oa_idx] if oa_idx >= 0 and oa_idx < len(headers) else "",
        "as_column": headers[as_idx] if as_idx >= 0 and as_idx < len(headers) else "",
    }


def _preview_rows(headers: List[str], rows: List[List[Any]], limit: int = 500) -> List[Dict[str, Any]]:
    out = []
    for row in rows[:limit]:
        out.append({header: row[idx] if idx < len(row) else "" for idx, header in enumerate(headers)})
    return out


def get_latest_file(utility: str, date_value: str) -> Dict[str, Any]:
    utility_name = _safe_segment(utility)
    date_folder = _safe_segment(date_value)
    transport, sftp, config = _connect_sftp()
    try:
        folder = _join_remote(config["base_dir"], utility_name, date_folder)
        files = _list_supported_files(sftp, folder)
        if not files:
            return {"utility": utility_name, "date": date_folder, "folder": folder, "file": None, "columns": [], "rows": [], "scan": []}
        latest = files[0]
        data = _read_remote_bytes(sftp, latest.path)
        headers, parsed_rows = _parse_file_rows(latest.name, data)
        return {
            "utility": utility_name,
            "date": date_folder,
            "folder": folder,
            "file": {
                "name": latest.name,
                "path": latest.path,
                "modified_at": latest.modified_at.isoformat() if latest.modified_at else None,
                "size": latest.size,
            },
            "columns": headers,
            "rows": _preview_rows(headers, parsed_rows),
            "row_count": len(parsed_rows),
            "scan": [
                {
                    "name": item.name,
                    "path": item.path,
                    "modified_at": item.modified_at.isoformat() if item.modified_at else None,
                    "size": item.size,
                }
                for item in files[:100]
            ],
            "total": _compute_total_card(headers, parsed_rows, latest.name),
        }
    finally:
        sftp.close()
        transport.close()


def _rows_to_csv_bytes(headers: List[str], rows: List[List[Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def _date_range(start: str, end: str) -> List[str]:
    start_date = datetime.strptime(start, "%Y-%m-%d").date()
    end_date = datetime.strptime(end, "%Y-%m-%d").date()
    if end_date < start_date:
        raise UtilityFileServiceError("to date must be after from date")
    if (end_date - start_date).days > 62:
        raise UtilityFileServiceError("Date range cannot exceed 63 days")
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    return dates


def export_range_zip(utility: str, from_date: str, to_date: str) -> Tuple[str, bytes]:
    utility_name = _safe_segment(utility)
    dates = _date_range(from_date, to_date)
    transport, sftp, config = _connect_sftp()
    try:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for day in dates:
                folder = _join_remote(config["base_dir"], utility_name, day)
                try:
                    files = _list_supported_files(sftp, folder)
                except Exception:
                    continue
                if not files:
                    continue
                latest = files[0]
                data = _read_remote_bytes(sftp, latest.path)
                headers, rows = _parse_file_rows(latest.name, data)
                csv_name = re.sub(r"[^A-Za-z0-9._-]+", "_", f"{day}_{latest.name}")
                csv_name = re.sub(r"\.(xlsx|xls|csv)$", ".csv", csv_name, flags=re.IGNORECASE)
                zf.writestr(csv_name, _rows_to_csv_bytes(headers, rows))
        zip_name = f"WBES_Portal_{re.sub(r'[^A-Za-z0-9._-]+', '_', utility_name)}_{from_date}_to_{to_date}.zip"
        return zip_name, zip_buf.getvalue()
    finally:
        sftp.close()
        transport.close()
