import os
import sys
import unittest
from io import BytesIO

from openpyxl import load_workbook

BACKEND_DIR = os.path.dirname(os.path.dirname(__file__))
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from services.sldc_attachment_converter import convert_telangana_csv_to_sldc_xlsx_bytes


class TelanganaSldcAttachmentConverterTests(unittest.TestCase):
    def test_dayahead_blank_station_schedule_blocks_are_zero(self):
        rows = [
            ["Name of Generator", "Singareni"],
            ["Plant name", "Singareni Collieries Company Limited-Sitarampatnam"],
            ["Capacity(MW)", "37"],
            ["Date", "2026-07-03"],
            ["Type", "dayahead"],
            [],
            [],
            ["Contract Type", "", "", "", "", "Lta"],
            ["Approval No", "", "", "", "", "TGTRANSCO/17/2024-25"],
            ["To Utility", "", "", "", "", "Utility"],
            ["Path", "", "", "", "", ""],
            ["Block", "Time Period", "Forecast(MW)", "AvC(MW)", "Station Schedule", "37"],
        ]
        for block in range(1, 97):
            station = "" if block <= 11 else "0"
            rows.append([block, f"2026-07-03 00:00:00", "", "0", station, station])

        csv_text = "\n".join(",".join(str(cell) for cell in row) for row in rows)
        content = convert_telangana_csv_to_sldc_xlsx_bytes(
            csv_text,
            plant_code="KOTHAGUDEM",
            fill_blank_avc=True,
            report_date="2026-07-02",
        )

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.worksheets[0]

        for excel_row in range(13, 24):
            self.assertEqual(ws.cell(row=excel_row, column=5).value, 0)


if __name__ == "__main__":
    unittest.main()
