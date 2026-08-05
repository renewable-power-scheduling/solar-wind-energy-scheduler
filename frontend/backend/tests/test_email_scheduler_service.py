import os
import csv
import io
import sys
import unittest
from io import BytesIO


BACKEND_DIR = os.path.dirname(os.path.dirname(__file__))
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from services.email_scheduler_service import load_email_scheduler_metadata, normalize_day_ahead_body
from services.sldc_attachment_converter import maybe_convert_for_auto_email
from services.sldc_attachment_converter import convert_ilios_pv_intraday_files_to_xlsx_bytes
from services.email_dispatch_service import build_email_html, build_email_plain_text
import main


class EmailSchedulerAttachmentNameTests(unittest.TestCase):
    def test_telangana_dsm_payload_does_not_fallback_to_zero_rows_without_s3_data(self):
        original_bucket = main._derive_s3_bucket_name
        try:
            main._derive_s3_bucket_name = lambda: ""

            payload = main._email_scheduler_build_dsm_payload_from_s3_for_email(
                plant_code="TELANGANA",
                report_date="2026-07-22",
            )

            self.assertIsNone(payload)
        finally:
            main._derive_s3_bucket_name = original_bucket

    def test_telangana_support_attachment_uses_computed_detail_values(self):
        payload = {
            "variant": "multi",
            "columns": [
                "DATE",
                "TO",
                "MONTH",
                "PROJECT",
                "INSTALLED CAPACITY (MW)",
                "GENERATION (KWH)",
                "DSM PENALTY (RS.), AS PER SCADA AVAILABILITY",
                "DSM PENALTY (RS.), AS MAINTENANCE INFORMATION",
                "PAISA/KWH SCADA AVAILABILITY",
                "PAISA/KWH MAINTENANCE INFORMATION",
                "SCADA AVAILABILITY(%)",
            ],
            "rows": [
                {
                    "DATE": "2026-07-22",
                    "TO": "2026-07-22",
                    "MONTH": "Jul-26",
                    "PROJECT": "KASIPET",
                    "INSTALLED CAPACITY (MW)": "15",
                    "GENERATION (KWH)": "625",
                    "DSM PENALTY (RS.), AS PER SCADA AVAILABILITY": "12",
                    "DSM PENALTY (RS.), AS MAINTENANCE INFORMATION": "12",
                    "PAISA/KWH SCADA AVAILABILITY": "1.92",
                    "PAISA/KWH MAINTENANCE INFORMATION": "1.92",
                    "SCADA AVAILABILITY(%)": "100%",
                    "__support_details": [
                        {
                            "block": 1,
                            "Schedule(Kwh)": 500,
                            "Meter data(KWh)": 625,
                            "AvC(Kwh)": 3750,
                            "Maintenance Update": 0,
                        }
                    ],
                }
            ],
        }

        attachment = main._email_scheduler_dsm_support_attachment_from_payload(
            payload=payload,
            plant_code="TELANGANA",
            report_date="2026-07-22",
        )

        self.assertIsNotNone(attachment)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(attachment["bytes"]), data_only=False)
        sheet = workbook["Kasipet"]
        row_idx = 10 + ((22 - 1) * 96)
        self.assertEqual(sheet.cell(row=row_idx, column=2).value, 500)
        self.assertEqual(sheet.cell(row=row_idx, column=3).value, 625)
        self.assertEqual(sheet.cell(row=row_idx, column=4).value, 3750)

    def test_da0_attachment_name_uses_next_day_date(self):
        name = main._email_scheduler_attachment_display_name(
            plant_code="KOTHAGUDEM",
            template_id="kothagudem_da0",
            schedule_type="dayahead",
            source_key="schedule_from_22.csv",
            original_name="KOTHAGUDEM_DA0.csv",
            report_date="2026-07-22",
        )

        self.assertEqual(name, "KOTHAGUDEM_DA0_23-07-2026.csv")

    def test_da1_attachment_name_uses_next_day_date(self):
        name = main._email_scheduler_attachment_display_name(
            plant_code="KOTHAGUDEM",
            template_id="kothagudem_da1",
            schedule_type="dayahead",
            source_key="schedule_from_88.csv",
            original_name="KOTHAGUDEM_DA1.xlsx",
            report_date="2026-07-22",
        )

        self.assertEqual(name, "KOTHAGUDEM_DA1_23-07-2026.xlsx")

    def test_template_screen_dayahead_attachment_keeps_supplied_schedule_date(self):
        name = main._email_scheduler_attachment_display_name(
            plant_code="KOTHAGUDEM",
            template_id="kothagudem_da0",
            schedule_type="dayahead",
            source_key="KOTHAGUDEM_DA0.csv",
            original_name="KOTHAGUDEM_DA0.csv",
            report_date="2026-07-23",
            date_already_day_ahead=True,
        )

        self.assertEqual(name, "KOTHAGUDEM_DA0_23-07-2026.csv")

    def test_sirmour_intraday_attachment_name_uses_current_date(self):
        name = main._email_scheduler_attachment_display_name(
            plant_code="SIRMOUR",
            template_id="sirmour_intraday",
            schedule_type="intraday",
            source_key="schedule_from_10.csv",
            original_name="Final_Schedule-Sirmour.xlsx",
            report_date="2026-07-22",
        )

        self.assertEqual(name, "Final_Schedule-Sirmour_22-07-2026.xlsx")

    def test_ilios_pv_intraday_combined_xlsx_uses_seven_site_format(self):
        site_files = {}
        for site in ["ANDAD", "ANJANGAON", "GUGARIYAKHEDI", "BALAKWADA", "BAMKHAL", "NANDGAON", "SAWDA"]:
            csv_text = "Block,Block Interval,Availability,Forecast\n"
            csv_text += "\n".join(
                f"{block},00:00-00:15,7.5,{1 if block == 44 else 0}"
                for block in range(1, 97)
            )
            site_files[site] = (f"{site}_schedule_from_9.csv", csv_text.encode("utf-8"))

        workbook_bytes = convert_ilios_pv_intraday_files_to_xlsx_bytes(
            site_files,
            report_date="2026-08-01",
            revision="9",
        )

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        sheet = workbook["REG"]
        self.assertEqual(sheet.max_column, 16)
        self.assertEqual(sheet.cell(row=1, column=1).value, "TYPE:")
        self.assertEqual(sheet.cell(row=2, column=2).value, "2026-08-01")
        self.assertEqual(sheet.cell(row=3, column=2).value, 9)
        self.assertEqual(sheet.cell(row=5, column=3).value, "M/s Physis Solar One Pvt Ltd Andad")
        self.assertEqual(sheet.cell(row=5, column=15).value, "M/s Physis Solar Power Two Pvt Ltd (SAWDA)")
        self.assertEqual(sheet.cell(row=6, column=16).value, "Forecast")
        self.assertEqual(sheet.cell(row=50, column=4).value, 1)
        self.assertEqual(sheet.cell(row=50, column=16).value, 1)


class EmailSchedulerBodyTests(unittest.TestCase):
    def test_da0_uses_day_ahead_zero(self):
        body = (
            "Dear Sir/Mam,\n"
            "Please find attached KOTHAGUDEM (37 MW) Day Ahead-01 Schedule for Date 11.06.2026."
        )
        self.assertEqual(
            normalize_day_ahead_body(body, "kothagudem_da0"),
            (
                "Dear Sir/Mam,\n"
                "Please find attached KOTHAGUDEM (37 MW) Day Ahead-0 Schedule for Date 11.06.2026."
            ),
        )

    def test_department_footer_is_appended_to_email_bodies(self):
        html_body = build_email_html(body_text="Dear Sir,\nPlease find attached.", employee_name="Ashwini Malkar")
        plain_body = build_email_plain_text(body_text="Dear Sir,\nPlease find attached.", employee_name="Ashwini Malkar")

        self.assertIn("Regards", html_body)
        self.assertIn("Ashwini Malkar", html_body)
        self.assertIn("Mob.: +91 8329261015", html_body)
        self.assertIn("Ashwini Malkar\nMob.: +91 8329261015", plain_body)
        self.assertIn("Forecasting And Scheduling Dept.", html_body)
        self.assertIn("forecasting.india@vedanjay-power.com", html_body)
        self.assertIn("http://www.vedanjay-power.com", html_body)
        self.assertIn("Forecasting And Scheduling Dept.", plain_body)
        self.assertIn("Mob.: +91 7666901814", plain_body)

    def test_auto_email_signature_can_skip_employee_mobile(self):
        html_body = build_email_html(
            body_text="Dear Sir,\nPlease find attached.",
            employee_name="Ashwini Malkar",
            include_employee_mobile=False,
        )
        plain_body = build_email_plain_text(
            body_text="Dear Sir,\nPlease find attached.",
            employee_name="Ashwini Malkar",
            include_employee_mobile=False,
        )

        self.assertIn("Ashwini Malkar", html_body)
        self.assertNotIn("Mob.: +91 8329261015", html_body)
        self.assertIn("Ashwini Malkar", plain_body)
        self.assertNotIn("Ashwini Malkar\nMob.: +91 8329261015", plain_body)

    def test_dsm_preview_is_inserted_before_regards(self):
        html_body = build_email_html(
            body_text="Dear Sir/Mam,\nPFA The DSM Penalty Report.",
            employee_name="Vedanjay Team",
            dsm_payload={
                "columns": ["Project", "DSM Penalty (Rs.)"],
                "rows": [{"Project": "SIRMOUR", "DSM Penalty (Rs.)": "718"}],
            },
        )

        body_idx = html_body.index("PFA The DSM Penalty Report.")
        preview_idx = html_body.index("DSM Report Preview")
        regards_idx = html_body.index("Regards")
        footer_idx = html_body.index("Forecasting And Scheduling Dept.")

        self.assertLess(body_idx, preview_idx)
        self.assertLess(preview_idx, regards_idx)
        self.assertLess(regards_idx, footer_idx)

    def test_da1_uses_day_ahead_one(self):
        body = (
            "Dear Sir/Mam,\n"
            "Please find attached KOTHAGUDEM (37 MW) Day Ahead-02 Schedule for Date 11.06.2026."
        )
        self.assertEqual(
            normalize_day_ahead_body(body, "kothagudem_da1"),
            (
                "Dear Sir/Mam,\n"
                "Please find attached KOTHAGUDEM (37 MW) Day Ahead-1 Schedule for Date 11.06.2026."
            ),
        )

    def test_non_day_ahead_body_is_unchanged(self):
        body = "Dear Sir/Mam,\nPlease find attached DSM report."
        self.assertEqual(normalize_day_ahead_body(body, "kothagudem_dsm"), body)

    def test_da0_subject_builder_uses_next_day_date(self):
        subject = main._email_scheduler_build_report_subject(
            template_id="kothagudem_da0",
            plant_code="KOTHAGUDEM",
            report_date="2026-07-29",
        )
        self.assertEqual(subject, "Dayahead Schedule KOTHAGUDEM (37 MW) for 30-07-2026")

    def test_da0_supplied_subject_date_is_shifted_for_send_now_and_schedule(self):
        subject = main._email_scheduler_subject_day_ahead_date(
            "Dayahead Schedule KOTHAGUDEM (37 MW) for 29-07-2026",
            template_id="kothagudem_da0",
            report_date="2026-07-29",
        )
        self.assertEqual(subject, "Dayahead Schedule KOTHAGUDEM (37 MW) for 30-07-2026")

    def test_da0_body_date_is_shifted_to_next_day(self):
        body = main._email_scheduler_day_ahead_body_date(
            "Dear Sir/Mam,\nPlease find attached KOTHAGUDEM (37 MW) Day Ahead-01 Schedule for Date 29.07.2026.",
            plant_code="KOTHAGUDEM",
            template_id="kothagudem_da0",
            report_date="2026-07-29",
        )
        self.assertEqual(
            body,
            "Dear Sir/Mam,\nPlease find attached KOTHAGUDEM (37 MW) Day Ahead-0 Schedule for Date 30.07.2026.",
        )

    def test_ilios_pv_email_metadata_is_active_with_dayahead_and_intraday(self):
        plants, templates_by_plant, _meta = load_email_scheduler_metadata()

        ilios = next((plant for plant in plants if plant.get("plant_code") == "ILIOS_PV"), None)
        self.assertIsNotNone(ilios)
        self.assertTrue(ilios.get("active"))

        templates = {item.get("id"): item for item in templates_by_plant.get("ILIOS_PV", [])}
        self.assertEqual(
            templates.get("ilios_pv_da0", {}).get("subject"),
            "Dayahead Schedule Ilios_PV (50MW) for {date_dashed}",
        )
        self.assertEqual(
            templates.get("ilios_pv_da0", {}).get("body"),
            "Dear Sir/Mam,\n\nPlease find attached Ilios_PV (50 MW) Day Ahead-Schedule for Date {date_dotted}",
        )
        self.assertEqual(templates.get("ilios_pv_intraday", {}).get("time_24h"), "17:00")
        self.assertEqual(
            templates.get("ilios_pv_intraday", {}).get("subject"),
            "Ilios_PV Intraday Schedule for the Month of {month_full}_{year_full}",
        )

    def test_ilios_pv_subject_and_body_builders_match_required_format(self):
        self.assertEqual(
            main._email_scheduler_build_report_subject(
                template_id="ilios_pv_da0",
                plant_code="ILIOS_PV",
                report_date="2026-08-01",
            ),
            "Dayahead Schedule Ilios_PV (50MW) for 02-08-2026",
        )
        self.assertEqual(
            main._email_scheduler_build_report_subject(
                template_id="ilios_pv_intraday",
                plant_code="ILIOS_PV",
                report_date="2026-08-01",
            ),
            "Ilios_PV Intraday Schedule for the Month of August_2026",
        )
        self.assertEqual(
            main._email_scheduler_ilios_pv_intraday_body("2026-08-01"),
            "Dear Sir/Mam,\n\nPlease find attached the Intraday Schedule ILIOS_PV for Date 01.08.2026",
        )


class EmailSchedulerAttachmentTests(unittest.TestCase):
    def test_osepl_dayahead_auto_email_uses_mh_vedanjay_da_template(self):
        lines = [
            "Name of Generator,OSEPL",
            "Plant name,OSEPL",
            "Capacity(MW),20",
            "Date,2026-07-03",
            "Type,dayahead",
            ",,,,,",
            ",,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,20",
        ]
        for block in range(1, 97):
            value = 0
            if 26 <= block <= 60:
                value = 2.5
            lines.append(f"{block},2026-07-03 00:00:00,,20,{value},20")

        converted = maybe_convert_for_auto_email(
            plant_code="OSEPL",
            template_id="osepl_da0",
            schedule_type="dayahead",
            file_name="schedule_from_22.csv",
            file_bytes="\n".join(lines).encode("utf-8"),
            report_date="2026-07-03",
        )

        self.assertIsNotNone(converted)
        self.assertTrue(converted.filename.endswith(".csv"))
        rows = list(csv.reader(io.StringIO(converted.content_bytes.decode("utf-8"))))

        self.assertEqual(rows[0], ["Schedule Template for MH_VEDANJAY and revision DA"])
        self.assertEqual(rows[1], ["", "Scheduling entity", "MH_VEDANJAY"])
        self.assertEqual(rows[2], ["", "Date", "2026-07-03"])
        self.assertEqual(rows[3], ["", "Revision No", "DA"])
        self.assertEqual(rows[5], ["POS Name", "Naldurg Inter 132kV", "Naldurg Inter 132kV", "Naldurg Inter 132kV"])
        self.assertEqual(rows[8], ["Contract ID", "", "", "CONTRACT00192"])
        self.assertEqual(rows[17], ["Capacity", "20", "20", "20"])
        self.assertEqual(rows[18], ["Block", "Declared Forecast", "Inter Avc", "Schedule"])
        self.assertEqual(len(rows[19:]), 96)

        block_25 = rows[19 + 24]
        block_26 = rows[19 + 25]
        block_60 = rows[19 + 59]
        block_61 = rows[19 + 60]
        self.assertEqual(block_25, ["25", "0", "0", "0"])
        self.assertEqual(block_26, ["26", "2.5", "20", "2.5"])
        self.assertEqual(block_60, ["60", "2.5", "20", "2.5"])
        self.assertEqual(block_61, ["61", "0", "0", "0"])
        for row in rows[19:]:
            self.assertEqual(row[1], row[3])

    def test_telangana_auto_email_xlsx_uses_template_header_metadata(self):
        lines = [
            "Name of Generator,timestamp",
            "Plant name,2026-07-02 00:00:00",
            "Capacity(MW),0",
            "Date,2026-07-02 00:30:00",
            "Type,2026-07-02 00:45:00",
            ",,,,,",
            ",,,,,",
            "Contract Type,,,,,",
            "Approval No,,,,,",
            "To Utility,,,,,",
            "Path,,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,10",
        ]
        for block in range(1, 97):
            lines.append(f"{block},2026-07-02 00:00:00,,10,0,0")

        expected_by_plant = {
            "BHUPALPALLY": (
                "Singareni Collieries Company Limited-Chelpur",
                10,
                "Mtoa",
                "TSTRANSCO/21/2023-24",
                "SCCL(BPL-003, BPL-006, BPL-028)",
            ),
            "KASIPET": (
                "Singareni Collieries Company Limited-Kasipet Mines",
                15,
                "Lta",
                "TSTRANSCO/20/2023-24",
                "SCCL(BPL-003, BPL-004, BPL-065)",
            ),
            "KOTHAGUDEM": (
                "Singareni Collieries Company Limited-Sitarampatnam",
                37,
                "Lta",
                "TGTRANSCO/17/2024-25",
                "General",
            ),
        }

        from openpyxl import load_workbook

        for selector in ("DA0", "DA1"):
            for plant_code, expected in expected_by_plant.items():
                with self.subTest(selector=selector, plant_code=plant_code):
                    converted = maybe_convert_for_auto_email(
                        plant_code=plant_code,
                        template_id=selector,
                        schedule_type=selector,
                        file_name="schedule_from_22.csv",
                        file_bytes="\n".join(lines).encode("utf-8"),
                        report_date="2026-07-01",
                    )

                    self.assertIsNotNone(converted)
                    workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
                    sheet = workbook.active
                    plant_name, capacity, contract_type, approval_no, to_utility = expected
                    self.assertEqual(sheet.cell(row=1, column=2).value, "Singareni")
                    self.assertEqual(sheet.cell(row=2, column=2).value, plant_name)
                    self.assertEqual(sheet.cell(row=3, column=2).value, capacity)
                    self.assertEqual(sheet.cell(row=4, column=2).value, "02-07-2026")
                    self.assertEqual(sheet.cell(row=5, column=2).value, "dayahead")
                    self.assertEqual(sheet.cell(row=8, column=6).value, contract_type)
                    self.assertEqual(sheet.cell(row=9, column=6).value, approval_no)
                    self.assertEqual(sheet.cell(row=10, column=6).value, to_utility)
                    self.assertEqual(sheet.cell(row=12, column=6).value, capacity)

    def test_telangana_da_selector_fills_blank_avc_in_auto_email_xlsx(self):
        capacities = {"KOTHAGUDEM": 37, "KASIPET": 15, "BHUPALPALLY": 10}
        from openpyxl import load_workbook

        for plant_code, capacity in capacities.items():
            for selector in ("DA0", "DA1"):
                with self.subTest(plant_code=plant_code, selector=selector):
                    lines = [
                        "Name of Generator,Singareni",
                        f"Plant name,{plant_code}",
                        f"Capacity(MW),{capacity}",
                        "Date,2026-06-30",
                        "Type,dayahead",
                        ",,,,,",
                        ",,,,,",
                        "Contract Type,,,,,Mtoa",
                        "Approval No,,,,,STRANSC0/21/2023-24",
                        "To Utility,,,,,BPL-003",
                        "Path,,,,,",
                        f"Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,{capacity}",
                    ]
                    for block in range(1, 97):
                        station = "0"
                        if 24 <= block <= 40:
                            station = "1.25"
                        lines.append(f"{block},2026-06-30 00:00:00,,,{station},{capacity}")

                    converted = maybe_convert_for_auto_email(
                        plant_code=plant_code,
                        template_id=selector,
                        schedule_type=selector,
                        file_name="schedule_from_22.csv",
                        file_bytes="\n".join(lines).encode("utf-8"),
                        report_date="2026-06-29",
                    )

                    self.assertIsNotNone(converted)
                    workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
                    sheet = workbook.active
                    self.assertEqual(sheet.cell(row=13, column=4).value, 0)
                    self.assertEqual(sheet.cell(row=36, column=4).value, capacity)
                    self.assertEqual(sheet.cell(row=52, column=4).value, capacity)
                    self.assertEqual(sheet.cell(row=53, column=4).value, 0)
                    self.assertEqual(sheet.cell(row=6, column=1).value, 0 if selector == "DA0" else 1)
                    self.assertEqual(sheet.cell(row=36, column=6).value, sheet.cell(row=36, column=5).value)
                    self.assertEqual(sheet.cell(row=13, column=6).value, sheet.cell(row=13, column=5).value)

    def test_telangana_auto_email_xlsx_always_writes_blocks_1_to_96(self):
        lines = [
            "Name of Generator,Singareni",
            "Plant name,Kothagudem",
            "Capacity(MW),37",
            "Date,2026-06-30",
            "Type,dayahead",
            ",,,,,",
            ",,,,,",
            "Contract Type,,,,,Mtoa",
            "Approval No,,,,,STRANSC0/21/2023-24",
            "To Utility,,,,,BPL-003",
            "Path,,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,37",
        ]
        source_blocks = list(range(12, 97)) + list(range(86, 97))
        self.assertEqual(len(source_blocks), 96)
        for block in source_blocks:
            lines.append(f"{block},2026-06-30 00:00:00,,37,1.25,1.25")

        converted = maybe_convert_for_auto_email(
            plant_code="KOTHAGUDEM",
            template_id="DA0",
            schedule_type="DA0",
            file_name="schedule_from_22.csv",
            file_bytes="\n".join(lines).encode("utf-8"),
            report_date="2026-06-29",
        )

        self.assertIsNotNone(converted)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
        sheet = workbook.active
        blocks = [int(sheet.cell(row=row, column=1).value) for row in range(13, 109)]
        self.assertEqual(blocks, list(range(1, 97)))

    def test_telangana_auto_email_xlsx_generates_full_day_time_periods(self):
        lines = [
            "timestamp,forecast",
        ]
        # Simulate a rotated/shifted DA source file. Time Period in the XLSX
        # must still be generated from block 1..96, not copied by row offset.
        source_blocks = list(range(12, 97)) + list(range(1, 12))
        self.assertEqual(len(source_blocks), 96)
        for idx, block in enumerate(source_blocks):
            hour = (idx * 15) // 60
            minute = (idx * 15) % 60
            lines.append(f"{block},2026-07-02 {hour:02d}:{minute:02d}:00,,10,1.25,1.25")

        converted = maybe_convert_for_auto_email(
            plant_code="BHUPALPALLY",
            template_id="DA0",
            schedule_type="DA0",
            file_name="schedule_from_22.csv",
            file_bytes="\n".join(lines).encode("utf-8"),
            report_date="2026-07-01",
        )

        self.assertIsNotNone(converted)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.cell(row=13, column=2).value, "2026-07-02 00:00:00")
        self.assertEqual(sheet.cell(row=14, column=2).value, "2026-07-02 00:15:00")
        self.assertEqual(sheet.cell(row=108, column=2).value, "2026-07-02 23:45:00")
        self.assertTrue(all(sheet.cell(row=row, column=2).value for row in range(13, 109)))

    def test_telangana_da0_da1_auto_email_avc_follows_station_schedule_window(self):
        base_lines = [
            "Name of Generator,Singareni",
            "Plant name,Bhupalpally",
            "Capacity(MW),10",
            "Date,2026-07-02",
            "Type,dayahead",
            ",,,,,",
            ",,,,,",
            "Contract Type,,,,,Mtoa",
            "Approval No,,,,,TSTRANSCO/21/2023-24",
            "To Utility,,,,,BPL-003",
            "Path,,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,10",
        ]
        source_blocks = list(range(12, 97)) + list(range(1, 12))
        self.assertEqual(len(source_blocks), 96)
        for block in source_blocks:
            station = "1.25" if 13 <= block <= 64 else "0"
            base_lines.append(f"{block},2026-07-02 00:00:00,,,{station},{station}")

        from openpyxl import load_workbook

        for selector in ("DA0", "DA1"):
            with self.subTest(selector=selector):
                converted = maybe_convert_for_auto_email(
                    plant_code="BHUPALPALLY",
                    template_id=selector,
                    schedule_type=selector,
                    file_name="schedule_from_22.csv",
                    file_bytes="\n".join(base_lines).encode("utf-8"),
                    report_date="2026-07-01",
                )

                self.assertIsNotNone(converted)
                workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
                sheet = workbook.active
                # Excel row = block + 12.
                self.assertEqual(sheet.cell(row=24, column=4).value, 0)
                self.assertEqual(sheet.cell(row=25, column=4).value, 10)
                self.assertEqual(sheet.cell(row=76, column=4).value, 10)
                self.assertEqual(sheet.cell(row=77, column=4).value, 0)

    def test_telangana_da1_generated_email_keeps_forecast_blank(self):
        lines = [
            "Name of Generator,Singareni",
            "Plant name,Bhupalpally",
            "Capacity(MW),10",
            "Date,2026-07-07",
            "Type,dayahead",
            ",,,,,",
            ",,,,,",
            "Contract Type,,,,,Mtoa",
            "Approval No,,,,,TSTRANSCO/21/2023-24",
            "To Utility,,,,,BPL-003",
            "Path,,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,10",
        ]
        for block in range(1, 97):
            lines.append(f"{block},2026-07-07 00:00:00,9.99,10,1.25,1.25")

        converted = maybe_convert_for_auto_email(
            plant_code="BHUPALPALLY",
            template_id="DA1",
            schedule_type="DA1",
            file_name="schedule_from_88.csv",
            file_bytes="\n".join(lines).encode("utf-8"),
            report_date="2026-07-07",
            source_key="generated/vedanjay/BHUPALPALLY/outputs/2026-07-07/Day-ahead/schedule_from_88.csv",
        )

        self.assertIsNotNone(converted)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
        sheet = workbook.active
        self.assertIsNone(sheet.cell(row=13, column=3).value)
        self.assertEqual(sheet.cell(row=13, column=5).value, 1.25)
        self.assertEqual(sheet.cell(row=13, column=6).value, 1.25)

    def test_telangana_da1_manual_email_copies_edited_value_to_forecast_station_and_helper(self):
        lines = ["block,mw"]
        for block in range(1, 97):
            value = 2.75 if block == 24 else 0
            lines.append(f"{block},{value}")

        converted = maybe_convert_for_auto_email(
            plant_code="BHUPALPALLY",
            template_id="DA1",
            schedule_type="DA1",
            file_name="edited_schedule.csv",
            file_bytes="\n".join(lines).encode("utf-8"),
            report_date="2026-07-07",
            source_key="manual-edits/vedanjay/BHUPALPALLY/2026-07-07/DA/manual-123-test/edited_schedule.csv",
        )

        self.assertIsNotNone(converted)
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.cell(row=36, column=3).value, 2.75)
        self.assertEqual(sheet.cell(row=36, column=5).value, 2.75)
        self.assertEqual(sheet.cell(row=36, column=6).value, 2.75)

    def test_telangana_day_ahead_email_forces_attachment_date_from_target_date(self):
        lines = [
            "Name of Generator,Singareni",
            "Plant name,Singareni Collieries Company Limited-Kasipet Mines",
            "Capacity(MW),15",
            "Date,10-07-2026",
            "Type,dayahead",
            ",,,,,",
            ",,,,,",
            "Contract Type,,,,,Lta",
            "Approval No,,,,,TSTRANSCO/20/2023-24",
            "To Utility,,,,,BPL-003",
            "Path,,,,,",
            "Block,Time Period,Forecast(MW),AvC(MW),Station Schedule,15",
        ]
        for block in range(1, 97):
            lines.append(f"{block},2026-07-10 00:00:00,9.99,15,1.25,1.25")

        from openpyxl import load_workbook

        for selector in ("DA0", "DA1"):
            converted = maybe_convert_for_auto_email(
                plant_code="KASIPET",
                template_id=selector,
                schedule_type=selector,
                file_name="schedule_from_88.csv" if selector == "DA1" else "schedule_from_22.csv",
                file_bytes="\n".join(lines).encode("utf-8"),
                report_date="2026-07-11",
                source_key=f"generated/vedanjay/KASIPET/outputs/2026-07-11/Day-ahead/schedule_from_88.csv",
            )

            self.assertIsNotNone(converted)
            workbook = load_workbook(BytesIO(converted.content_bytes), data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=4, column=2).value, "11-07-2026")
            self.assertEqual(sheet.cell(row=13, column=2).value, "2026-07-11 00:00:00")


if __name__ == "__main__":
    unittest.main()
